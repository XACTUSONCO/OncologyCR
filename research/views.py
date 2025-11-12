import collections, itertools
from datetime import datetime, date, timedelta
from urllib.parse import urlencode, parse_qs, quote
import json
import io, calendar
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

import xlsxwriter
from django.core import serializers

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods

from .models import Research, UploadFile, UploadEngFile, UploadInclusion, UploadExclusion, UploadReference, History, WaitingList, Cancer, research_WaitingList, Phase, Pre_Initiation, Pre_Initiation_IIT, Pre_Initiation_SIT, ONCO_CR_COUNT, \
                    End_research, UploadImage, End_UploadImage, Research_Archive, End_Research_Archive, Study_Category, Study_SubCategory, Study_Memo, DownloadLog, Line
from miscellaneous.models import Research_Management, Delivery, Supporting
from .utils import compare_research_fields, generate_search_query, generate_end_study_search_query
from feedback.utils import compare_assignment_fields
from feedback.models import Assignment, Feedback, STATUS_HISTORY, UploadRECIST
from dataroom.models import Image

from django.db.models import Value, Q, Count, F, Sum, Max, Min, ExpressionWrapper, DateField, DurationField, \
    IntegerField, FloatField, Case, When, Func, CharField, Subquery, OuterRef, Exists, Prefetch
from django.db.models.functions import TruncMonth, Cast, ExtractMonth, Coalesce, ExtractDay, TruncDate
from django.views.generic import ListView, View
from user.models import AuditEntry, Contact, User, InvestigatorContact, Team
from itertools import chain
from operator import itemgetter
import pandas as pd
import numpy as np
from dateutil import relativedelta
from django.core.exceptions import SuspiciousOperation
from administration.models import Company

TYPE_CHOICES = [
    ('IIT', 'IIT'),
    ('SIT', 'SIT'),
    ('EAP', 'EAP'),
    ('PMS', 'PMS'),
    ('완화 연구', '완화 연구'),
    ('etc', 'etc'),
]

@login_required()
def search_research(request):
    crc = Contact.objects.filter(Q(onco_A=1) & ~Q(team__name='etc')).order_by('name')

    if not request.GET or 'cancer' not in request.GET:        
        all_research, query_dict = generate_search_query(request)
        search_count = all_research.count()

        return render(
            request, 'pages/research/search.html', {
                #'waitinglist_field_choice': WaitingList.field_value_and_text(),
                'field_choice': Research.create_field_value_and_text(),
                'editable': True,
                'query': query_dict,
                'search': all_research,
                'search_count': search_count,
                'field_key_value': Research.create_field_value_and_text_dict(),
                #'field_key_id': Research.create_field_id_and_text_dict(),
                'crc': crc,
                'contact_userID_and_text': Research.contact_userID_and_text(),
            }
        )


@login_required()
def add_research(request):
    # GET req
    if request.method == 'GET':
        new_research = Research()
        backup = Contact.objects.filter(onco_A=1).order_by('name')
        companies = Company.objects.filter(is_deleted=0).order_by('name_kor')
        return render(
            request, 'pages/research/add.html', {
                'backup': backup,
                'companies': companies,
                'research': new_research,
                'editable': True,
                'field_choice': Research.create_field_value_and_text(),
                'contact_value_and_text': Research.contact_value_and_text(),
                'teams': Team.objects.all()}
        )

    temp_research, errors = Research.research_form_validation(request)

    if errors:
        # Temporary set attributes. cf. It is not a model instance!
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        backup = Contact.objects.filter(onco_A=1).order_by('name')
        companies = Company.objects.filter(is_deleted=0).order_by('name_kor')
        return render(
            request, 'pages/research/add.html',
            {'backup': backup,
             'companies': companies,   
             'research': temp_research,
             'editable': True,
             'field_choice': Research.create_field_value_and_text(),
             'contact_value_and_text': Research.contact_value_and_text(),
             'errors': error_msg,
             'teams': Team.objects.all()}
        )

    field_dict = dict(vars(temp_research))
    field_dict.pop('crc')
    field_dict.pop('cancer')
    field_dict.pop('phase')
    #field_dict.pop('alternation')
    field_dict.pop('line')
    #field_dict.pop('chemotherapy')
    field_dict.pop('type')
    field_dict.pop('route_of_administration')
    new_research = Research(**field_dict)
    new_research.save()

    new_research.crc.set(temp_research.crc)
    new_research.cancer.set(temp_research.cancer)
    new_research.phase.set(temp_research.phase)
    #new_research.alternation.set(temp_research.alternation)
    new_research.line.set(temp_research.line)
    #new_research.chemotherapy.set(temp_research.chemotherapy)
    new_research.type.set(temp_research.type)
    new_research.route_of_administration.set(temp_research.route_of_administration)

    # ONCO_CR_COUNT 객체 생성
    r_targets = request.POST.get('target')  # Assuming r_target is a multiple choice field
    for r_target in [r.strip() for r in r_targets.split(',') if r.strip()]:
        onco_cr_count = ONCO_CR_COUNT(research=new_research, r_target=r_target)
        onco_cr_count.save()

    field_summary = compare_research_fields({}, new_research.json())
    file_summary = collections.defaultdict(list)

     # 연구 프로토콜 파일 (국문)
    if request.FILES.getlist('file[]'):
        file_summary['prev'] = None

    for _file in request.FILES.getlist('file[]'):
        new_file = UploadFile(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 연구 프로토콜 파일 (영문)
    if request.FILES.getlist('eng_file[]'):
        file_summary['prev'] = None

    for _file in request.FILES.getlist('eng_file[]'):
        new_file = UploadEngFile(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 주 선정기준 이미지
    if request.FILES.get('inclusion'):
        file_summary['prev'] = None

    if request.FILES.get('inclusion'):
        _file = request.FILES.get('inclusion')
        new_file = UploadInclusion(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 주 제외기준 이미지
    if request.FILES.get('exclusion'):
        file_summary['prev'] = None

    if request.FILES.get('exclusion'):
        _file = request.FILES.get('exclusion')
        new_file = UploadExclusion(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # Reference
    if request.FILES.get('reference'):
        file_summary['prev'] = None

    if request.FILES.get('reference'):
        _file = request.FILES.get('reference')
        new_file = UploadReference(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 연구 바인더 위치 (사진)
    if request.FILES.getlist('image[]'):
        file_summary['prev'] = None

    for _image in request.FILES.getlist('image[]'):
        new_image = UploadImage(imagename=_image.name, image=_image, research=new_research)
        new_image.save()
        file_summary['curr'].append(new_image.json())

    # 장기 보관 문서 (파일)
    if request.FILES.getlist('research_archive[]'):
        file_summary['prev'] = None

    for _file in request.FILES.getlist('research_archive[]'):
        new_file = Research_Archive(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    history = History(user=request.user,
                      research=new_research,
                      summary={
                          'field_summary': field_summary,
                          'file_summary': file_summary
                      },
                      history_type=History.CREATE_RESEARCH,
                      content=new_research.json())
    history.save()

    return HttpResponseRedirect('/research/' + str(new_research.id) + '/')


@login_required()
@require_http_methods(['GET', 'POST'])
def edit_research(request, research_id):
    research = get_object_or_404(Research, pk=research_id)
    backup = Contact.objects.filter(onco_A=1).order_by('name')
    crc = Contact.objects.filter(Q(onco_A=1) & ~Q(team__name='etc')).order_by('name')
    companies = Company.objects.filter(is_deleted=0).order_by('name_kor')

    if research.is_deleted:
        return HttpResponse(status=403, content='이미 삭제된 연구입니다.')

    # if not request.user.is_superuser or request.user != research.uploader:
    #     return render(
    #         request, 'pages/research/not_allowed.html', status=403
    #     )

    # GET req
    if request.method == 'GET':
        return render(
            request, 'pages/research/edit.html',
            {'backup': backup,
             'crc': crc,
             'companies': companies,
             'research': research,
             'editable': True,
             'field_choice': Research.create_field_value_and_text(),
             'upload_files': UploadFile.objects.filter(is_deleted=False, research=research),
             'upload_engfiles': UploadEngFile.objects.filter(is_deleted=False, research=research),
             'upload_inclusions': UploadInclusion.objects.filter(is_deleted=False, research=research),
             'upload_exclusions': UploadExclusion.objects.filter(is_deleted=False, research=research),
             'upload_references': UploadReference.objects.filter(is_deleted=False, research=research),
             'upload_images': UploadImage.objects.filter(is_deleted=False, research=research),
             'research_archives': Research_Archive.objects.filter(is_deleted=False, research=research),
             'contact_value_and_text': Research.contact_value_and_text(),
             'teams': Team.objects.all(),
             'onco_cr_counts': research.onco_cr_count_set.all()
             }
        )

    temp_research, errors = Research.research_form_validation(request)
    backup = Contact.objects.filter(onco_A=1).order_by('name')
    crc = Contact.objects.filter(Q(onco_A=1) & ~Q(team__name='etc')).order_by('name')
    companies = Company.objects.filter(is_deleted=0).order_by('name_kor')

    if errors:
        # Temporary set attributes. cf. It is not a model instance!
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        temp_research.id = research.id
        return render(
            request, 'pages/research/edit.html',
            {
                'backup': backup,
                'crc': crc,
                'companies': companies,
                'research': temp_research,
                'editable': True,
                'field_choice': Research.create_field_value_and_text(),
                'errors': error_msg,
                'upload_files': UploadFile.objects.filter(is_deleted=False, research=research),
                'upload_engfiles': UploadEngFile.objects.filter(is_deleted=False, research=research),
                'upload_inclusions': UploadInclusion.objects.filter(is_deleted=False, research=research),
                'upload_exclusions': UploadExclusion.objects.filter(is_deleted=False, research=research),
                'upload_references': UploadReference.objects.filter(is_deleted=False, research=research),
                'upload_images': UploadImage.objects.filter(is_deleted=False, research=research),
                'research_archives': Research_Archive.objects.filter(is_deleted=False, research=research),
                'contact_value_and_text': Research.contact_value_and_text(),
                'teams': Team.objects.all(),
                'onco_cr_counts': research.onco_cr_count_set.all()
            }
        )

    prev_research_json = research.json()

    research.is_recruiting = temp_research.is_recruiting
    research.is_pre_screening = temp_research.is_pre_screening
    research.team = temp_research.team
    research.status = temp_research.status
    research.binder_location = temp_research.binder_location
    research.study_coordinator = temp_research.study_coordinator
    research.storage_date = temp_research.storage_date
    research.end_brief = temp_research.end_brief
    research.result_brief = temp_research.result_brief
    research.CRO = temp_research.CRO
    research.CRA_name = temp_research.CRA_name
    research.CRA_phoneNumber = temp_research.CRA_phoneNumber
    research.irb_number = temp_research.irb_number
    research.cris_number = temp_research.cris_number
    research.research_name = temp_research.research_name
    research.study_code = temp_research.study_code
    research.research_explanation = temp_research.research_explanation
    research.PI = temp_research.PI
    research.contact = temp_research.contact
    research.medicine_name = temp_research.medicine_name
    research.arm_name = temp_research.arm_name
    research.first_backup = temp_research.first_backup
    research.second_backup = temp_research.second_backup
    #research.lesion = temp_research.lesion
    #research.pdl1 = temp_research.pdl1
    #research.io_naive = temp_research.io_naive
    #research.brain_mets = temp_research.brain_mets
    #research.biopsy = temp_research.biopsy
    #research.turn_around_time = temp_research.turn_around_time
    #research.liver_function = temp_research.liver_function
    #research.lung_function = temp_research.lung_function
    #research.heart_function = temp_research.heart_function
    #research.kidney_function = temp_research.kidney_function
    research.remark = temp_research.remark

    research.crc.set(temp_research.crc)
    research.cancer.set(temp_research.cancer)
    research.phase.set(temp_research.phase)
    #research.alternation.set(temp_research.alternation)
    research.line.set(temp_research.line)
    #research.chemotherapy.set(temp_research.chemotherapy)
    research.type.set(temp_research.type)
    research.route_of_administration.set(temp_research.route_of_administration)
    research.save()

    curr_research_json = research.json()
    field_summary = compare_research_fields(prev_research_json,
                                            curr_research_json)
    file_summary = collections.defaultdict(list)

    # 연구 프로토콜 파일 (국문)
    if request.FILES.getlist('file[]'):
        prev_files = UploadFile.objects.filter(is_deleted=False, research=research)
        file_summary['prev'] = [f.json() for f in prev_files]
        prev_files.update(is_deleted=True)

    for _file in request.FILES.getlist('file[]'):
        new_file = UploadFile(filename=_file.name, file=_file, research=research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 연구 프로토콜 파일 (영문)
    if request.FILES.getlist('eng_file[]'):
        prev_files = UploadEngFile.objects.filter(is_deleted=False, research=research)
        file_summary['prev'] = [f.json() for f in prev_files]
        prev_files.update(is_deleted=True)

    for _file in request.FILES.getlist('eng_file[]'):
        new_file = UploadEngFile(filename=_file.name, file=_file, research=research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 주 선정기준 이미지
    if request.FILES.get('inclusion'):
        prev_files = UploadInclusion.objects.filter(is_deleted=False, research=research)
        file_summary['prev'] = [f.json() for f in prev_files]
        prev_files.update(is_deleted=True)

    if request.FILES.get('inclusion'):
        _file = request.FILES.get('inclusion')
        new_file = UploadInclusion(filename=_file.name, file=_file, research=research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 주 제외기준 이미지
    if request.FILES.get('exclusion'):
        prev_files = UploadExclusion.objects.filter(is_deleted=False, research=research)
        file_summary['prev'] = [f.json() for f in prev_files]
        prev_files.update(is_deleted=True)

    if request.FILES.get('exclusion'):
        _file = request.FILES.get('exclusion')
        new_file = UploadExclusion(filename=_file.name, file=_file, research=research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # Reference
    if request.FILES.get('reference'):
        prev_files = UploadReference.objects.filter(is_deleted=False, research=research)
        file_summary['prev'] = [f.json() for f in prev_files]
        prev_files.update(is_deleted=True)

    if request.FILES.get('reference'):
        _file = request.FILES.get('reference')
        new_file = UploadReference(filename=_file.name, file=_file, research=research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 연구 바인더 위치 (사진)
    if request.FILES.getlist('image[]'):
        prev_images = UploadImage.objects.filter(is_deleted=False, research=research)
        file_summary['prev'] = [i.json() for i in prev_images]
        prev_images.update(is_deleted=False)

    for _image in request.FILES.getlist('image[]'):
        new_image = UploadImage(imagename=_image.name, image=_image, research=research)
        new_image.save()
        file_summary['curr'].append(new_image.json())

    # 장기 보관 문서 (파일)
    if request.FILES.getlist('research_archive[]'):
        prev_files = Research_Archive.objects.filter(is_deleted=False, research=research)
        file_summary['prev'] = [f.json() for f in prev_files]
        prev_files.update(is_deleted=True)

    for _file in request.FILES.getlist('research_archive[]'):
        new_file = Research_Archive(filename=_file.name, file=_file, research=research)
        new_file.save()
        file_summary['curr'].append(new_file.json())
    
    if len(request.FILES.getlist('research_archive[]')) > 0:
        binder_image = UploadImage.objects.filter(is_deleted=False, research=research)
        binder_image.update(is_deleted=True)

    history = History(user=request.user, research=research,
                      history_type=History.EDIT_RESEARCH,
                      summary={
                          'field_summary': field_summary,
                          'file_summary': file_summary
                      },
                      content=research.json())
    history.save()

    return HttpResponseRedirect('/research/' + str(research.id) + '/')


@login_required()
def delete_research(request, research_id):
    research = Research.objects.get(pk=research_id)
    research.is_deleted = True
    research.save()

    history = History(user=request.user,
                      research=research,
                      summary={
                          'research_summary': research.research_name +
                                              ' (' + research.medicine_name + ') 삭제'},
                      history_type=History.REMOVE_RESEARCH,
                      content=research.json())
    history.save()

    prev_url = request.META.get('HTTP_REFERER')
    if '/research/search' in prev_url:
        return HttpResponseRedirect(prev_url)
    else:
        return HttpResponseRedirect('/research/search/')


@login_required()
def detail_research(request, research_id):
    research = get_object_or_404(
        Research.objects.prefetch_related('crc', 'history_set__user', 'history_set__research'),
        pk=research_id
    )

    query, query_dict = generate_search_query(request)
    #research_management = Research_Management.objects.filter(is_deleted=0)
    #research_management_field_key_value = Research_Management.create_field_value_and_text_dict
    this_monday = date.today() - timedelta(days=date.today().weekday())
    this_sunday = this_monday + timedelta(days=6)
    delivery_list = Delivery.objects.filter(is_deleted=0, visit_date__gte=this_monday, visit_date__lte=this_sunday)
    delivery_assignments = Assignment.objects.filter(is_deleted='0', research_id=94).order_by('create_date')

    assignments = research.assignment_set.filter(
        Q(is_deleted=False) & ~Q(status__in=['pre-screening', 'pre-screening-fail'])
    ).select_related(
        'research',
        'curr_crc'
    ).prefetch_related(
        'research__onco_cr_count_set',
        Prefetch(
            'feedback_set',
            queryset=Feedback.objects.order_by('-id').prefetch_related('fu'),
            to_attr='prefetched_feedback'
        )
    )

    research_pre_screening_lists = list(
        research.assignment_set.filter(
            Q(is_deleted=False, status__in=['pre-screening', 'pre-screening-fail'])
        ).select_related(
            'research'
        ).prefetch_related(
            'research__onco_cr_count_set',
            Prefetch(
                'feedback_set',
                queryset=Feedback.objects.order_by('-id'),  # 최신순
                to_attr='feedbacks_list'
            )
        )
    )

    if request.method == 'GET':
        return render(
            request, 'pages/research/detail.html',
            {
             'research': research,
             'assignments': assignments,
             'research_pre_screening_lists': research_pre_screening_lists,
             'research_waiting_list': research.research_waitinglist_set.filter(is_deleted=False).order_by('create_date'),
                
             'waitinglist_field_choice': research_WaitingList.field_value_and_text(),
             'history_key_value': History.INV_CHOICES,
             'query': query_dict,
             'editable': True,
             'field_choice': Research.create_field_value_and_text(),
             'field_key_value': Research.create_field_value_and_text_dict(),
             'upload_files': UploadFile.objects.filter(is_deleted=False, research=research),
             'upload_engfiles': UploadEngFile.objects.filter(is_deleted=False, research=research),
             'upload_inclusions': UploadInclusion.objects.filter(is_deleted=False, research=research),
             'upload_exclusions': UploadExclusion.objects.filter(is_deleted=False, research=research),
             'upload_references': UploadReference.objects.filter(is_deleted=False, research=research),
             'upload_images': UploadImage.objects.filter(is_deleted=False, research=research),
             'research_archives': Research_Archive.objects.filter(is_deleted=False, research=research),
             'upload_RECIST': UploadRECIST.objects.filter(is_deleted=False, assignment__in=research.assignment_set.filter(is_deleted=False).values('id')),

             #'research_management': research_management,
             #'research_management_field_key_value': research_management_field_key_value,
             'delivery_list': delivery_list,
             'delivery_assignments': delivery_assignments,
             'this_monday': this_monday,
             'this_sunday': this_sunday,
             'onco_cr_counts': research.onco_cr_count_set.all()
            }
        )


@login_required()
def update_target(request, research_id):
    if request.method == 'GET':
        return render(request, 'pages/research/partials/update_target.html',
                      {
                          'research': Research.objects.get(id=research_id),
                          'onco_cr_counts': ONCO_CR_COUNT.objects.filter(research=research_id)
                      })

    elif request.method == 'POST':
        r_target = request.POST.get('r_target')
        onco_cr_count_id = request.POST.get('onco_cr_count_id')
        is_update = request.POST.get('is_update')
        original_target = request.POST.get('original_target')
        if is_update == 'true':
            updated_onco_cr_count = ONCO_CR_COUNT.objects.get(pk=onco_cr_count_id)
            updated_onco_cr_count.r_target = r_target
            updated_onco_cr_count.save()
            Assignment.objects.filter(research_id=research_id, phase=original_target).update(phase=r_target)
        else:
            new_onco_cr_count = ONCO_CR_COUNT(research_id=research_id, r_target=r_target)
            new_onco_cr_count.save()

        return HttpResponseRedirect(f'/research/{research_id}/update_target/')
    else:
        raise SuspiciousOperation("Invalid request")


@login_required()
def delete_target(request):
    onco_cr_count_ids = request.POST.get('onco_cr_count_ids')
    ONCO_CR_COUNT.objects.filter(id__in=onco_cr_count_ids.split(',')).delete()
    return HttpResponse()


@login_required()
@require_http_methods(['GET', 'POST'])
def new_assignment(request, research_id):
    research = get_object_or_404(Research, pk=research_id)
    backup = Contact.objects.filter(onco_A=1).order_by('name')
    if request.method == 'GET':
        # Start creating new assignment
        return render(request, 'pages/assignment/add.html',
                      {
                          'research': research,
                          'backup': backup,
                          'assignment_field_choice': Assignment.field_value_and_text(),
                      })
    temp_assignment, errors = Assignment.assignment_form_validation(request, research)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/assignment/add.html',
            {
                'research': research,
                'backup': backup,
                'assignment': temp_assignment,
                'editable': True,
                'assignment_field_choice': Assignment.field_value_and_text(),
                'errors': error_msg
            }
        )

    # Create new assignment
    assignment = Assignment(**dict(vars(temp_assignment)))
    assignment.save()
    # Redirect to the new assignment detail page.

    field_summary = compare_assignment_fields({}, assignment.json())

    history = STATUS_HISTORY(user=request.user, assignment=assignment,
                      summary={
                          'field_summary': field_summary
                      },
                      history_type=STATUS_HISTORY.ADD_STATUS,
                      content=assignment.json())
    history.save()
    return HttpResponseRedirect('/assignment/' + str(assignment.id) + '/')

# Waiting List -> Screening
@login_required()
@require_http_methods(['GET', 'POST'])
def update_assignment(request, research_id, research_waitinglist_id):
    research = get_object_or_404(Research, pk=research_id)
    backup = Contact.objects.filter(onco_A=1).order_by('name')

    research_waitinglist = get_object_or_404(research_WaitingList, pk=research_waitinglist_id)
    research_waitinglist.is_deleted = True
    research_waitinglist.save()

    if request.method == 'GET':

        # Start creating new assignment
        return render(request, 'pages/assignment/update_assignment.html',
                      {
                          'research': research,
                          'assignment_field_choice': Assignment.field_value_and_text(),
                          'research_waitinglist': research_waitinglist,
                          'backup': backup,
                      })
    temp_assignment, errors = Assignment.assignment_form_validation(request, research)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/assignment/update_assignment.html',
            {
                'research': research,
                'assignment': temp_assignment,
                'editable': True,
                'assignment_field_choice': Assignment.field_value_and_text(),
                'errors': error_msg,
                'research_waitinglist': research_waitinglist,
                'backup': backup
            }
        )

    # From waitinglist To assignment
    assignment = Assignment(**dict(vars(temp_assignment)))
    assignment.save()

    for _file in request.FILES.getlist('file[]'):
        new_file = UploadRECIST(filename=_file.name, file=_file, assignment=assignment)
        new_file.save()

    return HttpResponseRedirect('/assignment/' + str(assignment.id) + '/')


# Pre Screening List -> Screening
@login_required()
@require_http_methods(['GET', 'POST'])
def update_pre_scr_assignment(request, research_id, pre_scr_assignment_id):
    research = get_object_or_404(Research, pk=research_id)
    backup = Contact.objects.filter(onco_A=1).order_by('name')

    pre_scr_assignment = get_object_or_404(Assignment, pk=pre_scr_assignment_id)
    pre_scr_assignment.is_deleted = False
    pre_scr_assignment.save()

    if request.method == 'GET':
        # Start creating new assignment
        return render(request, 'pages/assignment/update_pre_assignment.html',
                      {
                          'research': research,
                          'assignment_field_choice': Assignment.field_value_and_text(),
                          'pre_scr_assignment': pre_scr_assignment,
                          'backup': backup
                      })
    temp_assignment, errors = Assignment.assignment_form_validation(request, research)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/assignment/update_pre_assignment.html',
            {
                'research': research,
                'assignment': temp_assignment,
                'editable': True,
                'assignment_field_choice': Assignment.field_value_and_text(),
                'errors': error_msg,
                'pre_scr_assignment': pre_scr_assignment,
                'backup': backup
            }
        )

    assignment = Assignment(**dict(vars(temp_assignment)))
    assignment.save()

    for _file in request.FILES.getlist('file[]'):
        new_file = UploadRECIST(filename=_file.name, file=_file, assignment=assignment)
        new_file.save()

    return HttpResponseRedirect('/assignment/' + str(assignment.id) + '/')


@login_required()
def download_search(request):
    #query, _ = generate_search_query(request)
    #query.order_by('id')

    #filename = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    #filename = 'research_list_' + filename

    #workbook = Workbook()
    #worksheet = workbook.active

    #worksheet.append(Research.get_field_name())

    #for research in query:
    #    worksheet.append(research.tolist())

    #response = HttpResponse(
    #    content=save_virtual_workbook(workbook),
    #    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #response['Content-Disposition'] = \
    #    'attachment; filename=' + filename + '.xlsx'

    research_download = DownloadLog(downloader=request.user, content='연구 목록')
    research_download.save()
    
    return response


@login_required()
def download_assignment(request, research_id):
    research = get_object_or_404(Research, pk=research_id)

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # TODO Create research header

    d_format = '%Y-%m-%d'
    title = workbook.add_format({'border': 1, 'bottom': 2, 'align': 'center', 'bold': 1,'bg_color': '#BDBDBD', 'font_size': 13})
    border_center = workbook.add_format({'align': 'center', 'border': 1, 'bold': 1})
    assignment_start = workbook.add_format({'top': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})
    assignment_start2 = workbook.add_format({'top': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'font_size': 7})
    feedback_start = workbook.add_format({'top': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})
    feedback_start2 = workbook.add_format({'top': 1, 'left': 1, 'right': 1, 'text_wrap': True, 'valign': 'vcenter'})
    feedback_end = workbook.add_format({'bottom': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})
    feedback_end2 = workbook.add_format({'bottom': 1, 'left': 1, 'right': 1, 'text_wrap': True, 'valign': 'vcenter'})
    feedback_mid = workbook.add_format({'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})
    feedback_mid2 = workbook.add_format({'left': 1, 'right': 1, 'text_wrap': True, 'valign': 'vcenter'})

    # Create Assignment header
    assignment_header = [["Status", None],
                         ["Target", None],
                         ["Subject No.", None],
                         ["Patient No.", 20],
                         ["Name", 10],
                         ["Sex", None],
                         ["Age", None],
                         ["Cancer", 20],
                         ["previous Tx.", 30],
                         ["Doctor", 8],
                         ["Cycle", 8],
                         ["Visit Date", 10],
                         ["Tx, Dose", 12],
                         ["Imaging Date", 10],
                         ["Response", 8],
                         ["AE", 40],
                         ["Comment", 40],
                         ["ICF date", 10],
                         ["Screening Fail Date", 10]]

    title_style = title
    worksheet.merge_range('A1:S1', research.research_name, title_style)

    row = 1
    col = 0
    for a_header, width in assignment_header:
        worksheet.write(row, col, a_header, border_center)
        if width is not None:
            worksheet.set_column(col, col, width)
        col += 1

    row += 1
    col = 0
    
    if research.id == 400 or research.id == 401:
        assignments = research.assignment_set.filter(Q(is_deleted=0) & ~Q(status='pre-screening') & ~Q(status='pre-screening-fail')).order_by('no')
    else:
        assignments = research.assignment_set.filter(is_deleted=False).order_by('no')
   
    for assignment in assignments:
        style = assignment_start
        style2 = assignment_start2
        worksheet.write(row, col, assignment.status, style)
        col += 1
        worksheet.write(row, col, assignment.phase, style)
        col += 1
        worksheet.write(row, col, assignment.no, style)
        col += 1
        worksheet.write(row, col, assignment.register_number, style)
        col += 1
        worksheet.write(row, col, assignment.name, style)
        col += 1
        worksheet.write(row, col, assignment.sex, style)
        col += 1
        worksheet.write(row, col, assignment.age, style)
        col += 1
        worksheet.write(row, col, assignment.dx, style)
        col += 1
        worksheet.write(row, col, assignment.previous_tx, style2)
        col += 1
        worksheet.write(row, col, assignment.PI, style)

        feedbacks = assignment.feedback_set.all()
        if len(feedbacks) == 0:
            row += 1

        for i, feedback in enumerate(feedbacks):
            if i == 0:
                style = feedback_start
                style2 = feedback_start2
            elif i == len(feedbacks) - 1:
                style = feedback_end
                style2 = feedback_end2
            else:
                style = feedback_mid
                style2 = feedback_mid2

            if i > 0:  # feedback 2개 이상이면
                col = 0
                for _ in range(10):
                    worksheet.write_blank(row, col, None, style)
                    col += 1
            else:
                col = 10

            if feedback.day is None or feedback.day == '':
                worksheet.write(row, col, str(feedback.cycle), style)
            else:
                worksheet.write(row, col, 'C' + str(feedback.cycle) + 'D' + str(feedback.day), style)
            col += 1

            if feedback.dosing_date is None:
                worksheet.write_blank(row, col, None, style)
            else:
                worksheet.write(row, col, datetime.strftime(feedback.dosing_date, d_format), style)
            col += 1

            worksheet.write(row, col, feedback.tx_dose, style)
            col += 1

            if feedback.photo_date is None:
                worksheet.write_blank(row, col, None, style)
            else:
                worksheet.write(row, col, datetime.strftime(feedback.photo_date, d_format), style)
            col += 1

            worksheet.write(row, col, str(feedback.response) + str(feedback.response_text), style)
            col += 1
            worksheet.write(row, col, feedback.toxicity, style2)
            col += 1
            worksheet.write(row, col, feedback.comment, style2)
            col += 1

            if feedback.ICF_date is None:
                worksheet.write_blank(row, col, None, style)
            else:
                worksheet.write(row, col, datetime.strftime(feedback.ICF_date, d_format), style)
            col += 1

            if feedback.scr_fail is None:
                worksheet.write_blank(row, col, None, style)
            else:
                worksheet.write(row, col, datetime.strftime(feedback.scr_fail, d_format), style)

            row += 1

        col = 0

    # Close the workbook before sending the data.
    workbook.close()

    # Rewind the buffer.
    output.seek(0)

    filename = research.research_name + '.xlsx'
    try:
        filename.encode('ascii')
        file_expr = 'filename="{}"'.format(filename)
    except UnicodeEncodeError:
        file_expr = "filename*=utf-8''{}".format(quote(filename))

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; ' + file_expr

    return response


class CrcListView(ListView):
    model = ONCO_CR_COUNT
    context_object_name = 'oncology'
    template_name = 'pages/research/crc_research_list.html'

    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    date_year = today.year
    date_month = today.month
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    EOT_assign = Feedback.objects.filter(assignment__is_deleted=0, cycle='EOT').values('assignment_id')
    MAX_assign_status = STATUS_HISTORY.objects.filter(assignment__is_deleted=0) \
                                              .values('assignment_id') \
                                              .annotate(Max('create_date')).values('create_date__max')
    EOS_assign = Feedback.objects.exclude(eos__isnull=True).values('assignment_id').distinct() ##### 2024.03.07 

    queryset = ONCO_CR_COUNT.objects.filter(Q(research__is_deleted=False) & ~Q(research__status__in=['종료보고완료', '결과보고완료', '장기보관완료'])) \
        .annotate(custom_order=Case(When(research__is_recruiting='Recruiting', then=Value(1)),
                                    When(research__is_recruiting='Holding', then=Value(2)),
                                    When(research__is_recruiting='Not yet recruiting', then=Value(3)),
                                    When(research__is_recruiting='Completed', then=Value(4)),
                                    output_field=IntegerField()),
                  ATEAM=Case(When(research__onco_A='1', then=Value(1)),
                             When(research__onco_A='0', then=Value(2)),
                             output_field=IntegerField()),
                  screening_total_filter=Count('research__assignment', distinct=True,
                                               filter=(Q(research__assignment__feedback__ICF_date__isnull=False) &
                                                       Q(research__assignment__is_deleted=0) &
                                                       ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']))),
                  pre_screening_filter=Count('research__assignment', distinct=True,
                                             filter=(Q(research__assignment__status='pre-screening') &
                                                     Q(research__assignment__is_deleted=0) &
                                                     Q(research__assignment__feedback__ICF_date__isnull=False))),
                  pre_screening_fail_filter=Count('research__assignment', distinct=True,
                                                  filter=(Q(research__assignment__status='pre-screening-fail') &
                                                          Q(research__assignment__is_deleted=0) &
                                                          Q(research__assignment__feedback__scr_fail__isnull=False))),
                  scr_fail_total_filter=Count('research__assignment', distinct=True,
                                              filter=(Q(research__assignment__feedback__scr_fail__isnull=False) &
                                                      Q(research__assignment__is_deleted=0) &
                                                      ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']))),
                  enroll_total_filter=Count('research__assignment', distinct=True,
                                            filter=(Q(research__assignment__feedback__cycle='1',
                                                      research__assignment__feedback__day='1',
                                                      research__assignment__is_deleted=0))),
                  screening_filter=Count('research__assignment', distinct=True,
                                         filter=(Q(research__assignment__feedback__ICF_date__isnull=False) &
                                                 Q(research__assignment__phase=F('r_target'),
                                                   research__assignment__curr_crc=F('research__crc'),
                                                   research__assignment__is_deleted=0) &
                                                 ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']))),
                  screening_month_filter=Count('research__assignment', distinct=True,
                                        filter=(Q(research__assignment__feedback__ICF_date__isnull=False) &
                                                Q(research__assignment__phase=F('r_target'),
                                                  research__assignment__curr_crc=F('research__crc'),
                                                  research__assignment__is_deleted=0) &
                                               ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']) &
                                                Q(research__assignment__feedback__ICF_date__range=[from_date, to_date]))),

                  scr_fail_month_filter=Count('research__assignment', distinct=True,
                                              filter=(Q(research__assignment__feedback__scr_fail__range=[from_date, to_date]) &
                                                      Q(research__assignment__phase=F('r_target'),
                                                        research__assignment__curr_crc=F('research__crc'),
                                                        research__assignment__is_deleted=0) &
                                                     ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']))),
                  scr_fail_filter=Count('research__assignment', distinct=True,
                                        filter=(Q(research__assignment__feedback__scr_fail__isnull=False) &
                                                Q(research__assignment__phase=F('r_target'),
                                                  research__assignment__curr_crc=F('research__crc'),
                                                  research__assignment__is_deleted=0) &
                                                ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']))),

                  ongoing_month_filter=Count('research__assignment', distinct=True,
                                             filter=(Q(research__assignment__phase=F('r_target'),
                                                       research__assignment__curr_crc=F('research__crc'),
                                                       research__assignment__is_deleted=0) &
                                                     ((Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__gte=from_date) &
                                                       Q(research__assignment__feedback__cycle='EOT', research__assignment__feedback__dosing_date__lt=to_date)) |
                                                     (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__range=[from_date, to_date]) &
                                                      ~Q(research__assignment__id__in=EOS_assign))))), ##### 2024.03.07
                  ongoing_filter=Count('research__assignment', distinct=True,
                                       filter=(Q(research__assignment__phase=F('r_target'),
                                                 research__assignment__curr_crc=F('research__crc'),
                                                 research__assignment__is_deleted=0) &
                                               ((Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__gte=from_date) &
                                                 Q(research__assignment__feedback__cycle='EOT', research__assignment__feedback__dosing_date__lt=to_date)) |
                                                (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__range=[from_date, to_date]) &
                                                 ~Q(research__assignment__id__in=EOS_assign)) | ##### 2024.03.07
                                                (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__lt=from_date) &
                                                 ~Q(research__assignment__id__in=EOT_assign) & ~Q(research__assignment__id__in=EOS_assign))))), ##### 2024.03.07
                  enroll_filter=Count('research__assignment', distinct=True,
                                      filter=(Q(research__assignment__feedback__cycle='1',
                                                research__assignment__feedback__day='1') &
                                              Q(research__assignment__phase=F('r_target'),
                                                research__assignment__curr_crc=F('research__crc'),
                                                research__assignment__is_deleted=0))),
                  FU_filter=Count('research__assignment', distinct=True,
                                  filter=Q(research__assignment__status_history__create_date__in=MAX_assign_status,
                                           research__assignment__status_history__summary__contains='FU',
                                           research__assignment__phase=F('r_target'), research__assignment__is_deleted=0))) \
        .order_by('ATEAM', 'custom_order', 'research__research_name') \
        .prefetch_related('research', 'research__crc', 'research__cancer')


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 월별 ('Screening/Ongoing/Enroll')
        user = self.request.user
        team_name = getattr(getattr(user, 'contact', None), 'team', None)
        team_name = getattr(team_name, 'name', None)

        team_filter = Q(team__name=team_name) if team_name else Q()
        # 팀 + onco_A 필터링된 Contact ID 리스트
        contact_ids = Contact.objects.filter(Q(onco_A=1) & team_filter).values_list('id', flat=True)

        today = datetime.today()
        from_date = datetime(today.year, today.month, 1)
        date_year = today.year
        date_month = today.month
        to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
        ongo_condition1 = Feedback.objects.filter(cycle='1', day='1', dosing_date__lt=from_date).values('assignment_id')
        EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values( 'assignment_id')
        ongo_condition2 = Feedback.objects.filter(cycle='1', day='1', dosing_date__gte=from_date).values('assignment_id')
        EOT_assign2 = Feedback.objects.filter(Q(cycle='EOT', dosing_date__lte=today) & Q(assignment_id__in=ongo_condition2)).values('assignment_id')
        EOS_assign = Feedback.objects.exclude(eos__isnull=True).values('assignment_id').distinct() ##### 2024.03.07
        
        # screening condition
        C1D1_assign = Feedback.objects.filter(Q(cycle='1', day='1', dosing_date__year=date_year, dosing_date__month=date_month)).values('assignment_id')
        screening_condition = Feedback.objects.filter(
            (Q(ICF_date__year=date_year) & Q(ICF_date__month=date_month) &
             ~Q(assignment__status='pre-screening') & ~Q(assignment__status='pre-screening-fail') & ~Q(assignment_id__in=C1D1_assign))
        ).values('assignment_id').distinct()

        # enroll_month condition
        enroll_month_condition = Feedback.objects.filter(
            Q(cycle='1', day='1', dosing_date__gte=from_date, dosing_date__lte=to_date)).values('assignment_id').distinct()

        # ongoing condition
        ongoing_condition = Feedback.objects.filter(
            (Q(cycle='1', day='1', dosing_date__gte=from_date) & Q(cycle='EOT', dosing_date__lt=to_date)) |
            (Q(cycle='1', day='1', dosing_date__gte=from_date, dosing_date__lte=to_date) & ~Q(assignment_id__in=EOT_assign2) & ~Q(assignment_id__in=EOS_assign)) |  ##### 2024.03.07
            #(Q(cycle='EOT', dosing_date__year=date_year) & Q(cycle='EOT', dosing_date__month=date_month)) |
            (Q(cycle='1', day='1', dosing_date__lt=from_date) & ~Q(assignment_id__in=EOT_assign) & ~Q(assignment_id__in=EOS_assign)) ##### 2024.03.07
        ).values('assignment_id').distinct()

        observe_and_PMS = Research.objects.filter(Q(type__value='EAP') | Q(type__value='PMS') | Q(type__value='Palliative') | Q(type__value='Blood') | Q(type__value='ETC')).values('id')
        charts = Assignment.objects.filter(research__is_deleted=0, is_deleted=0, curr_crc__in=contact_ids) \
            .values('curr_crc__name').distinct().order_by('curr_crc__name') \
            .filter(~Q(curr_crc__in=[91, 15])) \
            .annotate(custom_order=Case(When(curr_crc__team__name='GSI', then=Value(1)),
                                        When(curr_crc__team__name='CLUE', then=Value(2)),
                                        When(curr_crc__team__name='etc', then=Value(4)),
                                        output_field=IntegerField())).order_by('custom_order', 'curr_crc__name') \
            .annotate(screening_chart=Count('id', filter=Q(id__in=screening_condition)),
                      ongoing_chart=Count('id', filter=(Q(id__in=ongoing_condition) & ~Q(research_id__in=observe_and_PMS))),
                      PMS_ongoing_chart=Count('id', filter=(Q(id__in=ongoing_condition) & Q(research_id__in=observe_and_PMS))),
                      enroll_chart=Count('id', filter=Q(id__in=enroll_month_condition) & ~Q(research_id__in=observe_and_PMS)),
                      PMS_enroll_chart=Count('id', filter=Q(id__in=enroll_month_condition) & Q(research_id__in=observe_and_PMS))
                      ).values('curr_crc__user__first_name', 'curr_crc__user__id', 'screening_chart',
                               'ongoing_chart', 'PMS_ongoing_chart', 'enroll_chart', 'PMS_enroll_chart')\
            .order_by('custom_order', 'curr_crc__name')

        crc_list = [];
        scr_list = [];
        ongo_list = [];
        PMS_ongo_list = [];
        enroll_list = [];
        PMS_enroll_list = [];
        for chart in charts:
            crc_list.append({'id': str(chart['curr_crc__user__id']), 'name': str(chart['curr_crc__user__first_name'])})
            scr_list.append(str(chart['screening_chart']))
            ongo_list.append(str(chart['ongoing_chart']))
            PMS_ongo_list.append(str(chart['PMS_ongoing_chart']))
            enroll_list.append(str(chart['enroll_chart']))
            PMS_enroll_list.append(str(chart['PMS_enroll_chart']))

        context['crc_list'] = crc_list
        context['scr_list'] = scr_list
        context['ongo_list'] = ongo_list
        context['PMS_ongo_list'] = PMS_ongo_list
        context['enroll_list'] = enroll_list
        context['PMS_enroll_list'] = PMS_enroll_list

        # 월별 ('Input Delay Days')
        # # 날짜 조건
        # dosing_date_range_q = Q(dosing_date__range=[from_date, to_date])
        # icf_range_q = Q(ICF_date__range=[from_date, to_date])
        #
        # # 각 gap condition을 Q 객체로 대체
        # gap_q1 = Q(ICF_date__isnull=False) & icf_range_q  # ICF
        # gap_q2 = ~Q(cycle='1', day='1') & ~Q(cycle='') & Q(ICF_date__isnull=True) & dosing_date_range_q  # Cycle Visit
        # gap_q3 = Q(cycle='1', day='1') & ~Q(cycle__isnull=True) & dosing_date_range_q  # C1D1
        #
        # gap_charts = Feedback.objects.filter(
        #     assignment__is_deleted=0,
        #     uploader__contact__in=contact_ids
        # ).exclude(
        #     uploader=193
        # ).annotate(
        #     custom_order=Case(
        #         When(uploader__contact__team__name='GSI', then=Value(1)),
        #         When(uploader__contact__team__name='CLUE', then=Value(2)),
        #         When(uploader__contact__team__name='etc', then=Value(4)),
        #         output_field=IntegerField()
        #     )
        # ).annotate(
        #     C1D1_count=Count('id', filter=gap_q3),
        #     gap1_count=Count('id', filter=gap_q1),
        #     gap2_count=Count('id', filter=gap_q2),
        #     gap1_total=Sum(
        #         ExpressionWrapper(
        #             Cast(F('create_date'), DateField()) - F('ICF_date'),
        #             output_field=IntegerField()
        #         ),
        #         filter=gap_q1
        #     ),
        #     gap2_total=Sum(
        #         ExpressionWrapper(
        #             Cast(F('create_date'), DateField()) - F('dosing_date'),
        #             output_field=IntegerField()
        #         ),
        #         filter=gap_q2
        #     )
        # ).annotate(
        #     total_gap=(Coalesce(F('gap1_total'), 0) + Coalesce(F('gap2_total'), 0)) /
        #               (Coalesce(F('gap1_count'), 0) + Coalesce(F('gap2_count'), 0) + Coalesce(F('C1D1_count'), 0)),
        #     visit_count=(Coalesce(F('gap1_count'), 0) + Coalesce(F('gap2_count'), 0) + Coalesce(F('C1D1_count'), 0))
        # ).values('uploader_id__first_name', 'uploader_id', 'total_gap', 'visit_count'
        # ).order_by('custom_order', 'uploader_id__first_name')

        # ICF data input interval
        gap_condition_1 = Feedback.objects.filter(ICF_date__isnull=False, ICF_date__gte=from_date, ICF_date__lte=to_date)
        # cycle data input interval
        gap_condition_2 = Feedback.objects.filter(~Q(cycle='1', day='1') & ~Q(cycle='') & Q(ICF_date__isnull=True) &
            Q(dosing_date__gte=from_date) & Q(dosing_date__lte=to_date))
        # C1D1 data input interval
        gap_condition_3 = Feedback.objects.filter(Q(cycle='1', day='1') & Q(cycle__isnull=False) &
            Q(dosing_date__gte=from_date) & Q(dosing_date__lte=to_date))

        gap_charts = Feedback.objects.filter(assignment__is_deleted=0, uploader__contact__in=contact_ids)\
            .values('uploader').distinct().order_by('uploader') \
            .filter(~Q(uploader=193)) \
            .annotate(custom_order=Case(When(uploader__contact__team__name='GSI', then=Value(1)), When(uploader__contact__team__name='CLUE', then=Value(2)), When(uploader__contact__team__name='etc', then=Value(4)),
                                        output_field=IntegerField())).order_by('custom_order', 'uploader_id__first_name') \
            .annotate(C1D1_count=Count('id', filter=Q(id__in=gap_condition_3.values('id'))),
                      total_gap=(Coalesce(Sum(ExpressionWrapper(Cast(F('create_date'), DateField()) - F('ICF_date'),
                        output_field=IntegerField()) / 86400000000, filter=Q(id__in=gap_condition_1.values('id'))), 0) +
                                 Coalesce(Sum(ExpressionWrapper(Cast(F('create_date'), DateField()) - F('dosing_date'),
                        output_field=IntegerField()) / 86400000000, filter=Q(id__in=gap_condition_2.values('id'))), 0)) /
                                (Count('id', filter=Q(id__in=gap_condition_1.values('id'))) +
                                 Count('id', filter=Q(id__in=gap_condition_2.values('id'))) + F('C1D1_count')),
                      visit_count=(Count('id', filter=Q(id__in=gap_condition_1.values('id'))) +
                                   Count('id', filter=Q(id__in=gap_condition_2.values('id'))) + F('C1D1_count'))
                      ).values('uploader_id__first_name', 'uploader_id', 'total_gap', 'visit_count')

        gap_crc_list = [];
        gap_count_list = [];
        visit_count_list = [];
        #no_input_count_list = [];
        for gap in gap_charts:
            gap_crc_list.append({'id': str(gap['uploader_id']), 'name': str(gap['uploader_id__first_name'])})
            gap_count_list.append(str(gap['total_gap']))
            visit_count_list.append(str(gap['visit_count']))
            #no_input_count_list.append(str(gap['no_input_count']))

        context['gap_crc_list'] = gap_crc_list
        context['gap_count_list'] = gap_count_list
        context['visit_count_list'] = visit_count_list
        #context['no_input_count_list'] = no_input_count_list

        # 월별 ('screening/ongoing/enroll 수', 'Input Delay Days')
        if self.request.GET.get('from_month_date'):
            month_today = datetime.strptime(self.request.GET.get('from_month_date'), '%Y-%m-%d')
            from_month_date = datetime(month_today.year, month_today.month, 1)
            date_year = month_today.year
            date_month = month_today.month
            to_month_date = datetime(month_today.year, month_today.month, calendar.monthrange(month_today.year, month_today.month)[1])
            ongo_condition1 = Feedback.objects.filter(cycle='1', day='1', dosing_date__lt=from_month_date).values('assignment_id')
            EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values('assignment_id')
            ongo_condition2 = Feedback.objects.filter(cycle='1', day='1', dosing_date__gte=from_month_date).values('assignment_id')
            EOT_assign2 = Feedback.objects.filter(Q(cycle='EOT', dosing_date__lte=to_month_date) & Q(assignment_id__in=ongo_condition2)).values('assignment_id')

            # screening condition
            C1D1_assign = Feedback.objects.filter(Q(cycle='1', day='1', dosing_date__year=date_year, dosing_date__month=date_month)).values('assignment_id')
            screening_condition = Feedback.objects.filter(
                (Q(ICF_date__year=date_year) & Q(ICF_date__month=date_month) &
                 ~Q(assignment__status='pre-screening') & ~Q(assignment__status='pre-screening-fail') & ~Q(assignment_id__in=C1D1_assign))
            ).values('assignment_id').distinct()

            # enroll_month condition
            enroll_month_condition = Feedback.objects.filter(
                Q(cycle='1', day='1', dosing_date__gte=from_month_date, dosing_date__lte=to_month_date)
            ).values('assignment_id').distinct()

            # ongoing condition
            ongoing_condition = Feedback.objects.filter(
                (Q(cycle='1', day='1', dosing_date__gte=from_month_date) & Q(cycle='EOT', dosing_date__lt=to_month_date)) |
                (Q(cycle='1', day='1', dosing_date__gte=from_month_date, dosing_date__lte=to_month_date) & ~Q(assignment_id__in=EOT_assign2) & ~Q(assignment_id__in=EOS_assign)) |  ##### 2024.03.07
                # (Q(cycle='1', day='1', dosing_date__year=date_year) & Q(cycle='1', day='1', dosing_date__month=date_month)) |
                # (Q(cycle='EOT', dosing_date__year=date_year) & Q(cycle='EOT', dosing_date__month=date_month)) |
                (Q(cycle='1', day='1', dosing_date__lt=from_month_date) & ~Q(assignment_id__in=EOT_assign) & ~Q(assignment_id__in=EOS_assign)) ##### 2024.03.07
            ).values('assignment_id').distinct()

            observe_and_PMS = Research.objects.filter(Q(type__value='EAP') | Q(type__value='PMS') | Q(type__value='Palliative') | Q(type__value='Blood') | Q(type__value='ETC')).values('id')

            team_filter = Q(team__name=self.request.GET.get('team') ) if self.request.GET.get('team') else Q()
            contact_ids = Contact.objects.filter(Q(onco_A=1) & team_filter).values_list('id', flat=True)

            charts = Assignment.objects.filter(research__is_deleted=0, is_deleted=0, curr_crc__in=contact_ids)\
                            .values('curr_crc__name').distinct().order_by('curr_crc__name')\
                            .filter(~Q(curr_crc__in=[91, 15])) \
                            .annotate(custom_order=Case(When(curr_crc__team__name='GSI', then=Value(1)),
                                                        When(curr_crc__team__name='CLUE', then=Value(2)),
                                                        When(curr_crc__team__name='etc', then=Value(4)),
                                                        output_field=IntegerField())).order_by('custom_order', 'curr_crc__name') \
                            .annotate(screening_chart=Count('id', filter=Q(id__in=screening_condition)),
                                      ongoing_chart=Count('id', filter=(Q(id__in=ongoing_condition) & ~Q(research_id__in=observe_and_PMS))),
                                      PMS_ongoing_chart=Count('id', filter=(Q(id__in=ongoing_condition) & Q(research_id__in=observe_and_PMS))),
                                      enroll_chart=Count('id', filter=Q(id__in=enroll_month_condition) & ~Q(research_id__in=observe_and_PMS)),
                                      PMS_enroll_chart=Count('id', filter=Q(id__in=enroll_month_condition) & Q(research_id__in=observe_and_PMS))
                                     ).values('curr_crc__user__first_name', 'curr_crc__user__id', 'screening_chart', 'ongoing_chart',
                                              'PMS_ongoing_chart', 'enroll_chart', 'PMS_enroll_chart')

            crc_list = [];
            scr_list = [];
            ongo_list = [];
            PMS_ongo_list = [];
            enroll_list = [];
            PMS_enroll_list = [];
            for chart in charts:
                crc_list.append({'id': str(chart['curr_crc__user__id']), 'name': str(chart['curr_crc__user__first_name'])})
                scr_list.append(str(chart['screening_chart']))
                ongo_list.append(str(chart['ongoing_chart']))
                PMS_ongo_list.append(str(chart['PMS_ongoing_chart']))
                enroll_list.append(str(chart['enroll_chart']))
                PMS_enroll_list.append(str(chart['PMS_enroll_chart']))

            context['crc_list'] = crc_list
            context['scr_list'] = scr_list
            context['ongo_list'] = ongo_list
            context['PMS_ongo_list'] = PMS_ongo_list
            context['enroll_list'] = enroll_list
            context['PMS_enroll_list'] = PMS_enroll_list

            # # 날짜 조건
            # dosing_date_range_q = Q(dosing_date__range=[from_month_date, to_month_date])
            # icf_range_q = Q(ICF_date__range=[from_month_date, to_month_date])
            #
            # # 각 gap condition을 Q 객체로 대체
            # gap_q1 = Q(ICF_date__isnull=False) & icf_range_q  # ICF
            # gap_q2 = ~Q(cycle='1', day='1') & ~Q(cycle='') & Q(ICF_date__isnull=True) & dosing_date_range_q  # Cycle Visit
            # gap_q3 = Q(cycle='1', day='1') & ~Q(cycle__isnull=True) & dosing_date_range_q  # C1D1
            #
            # # 대상 uploader (onco A 사용자)
            # onco_A_user_ids = Contact.objects.filter(onco_A=1).values_list('user_id', flat=True)
            #
            # gap_charts = Feedback.objects.filter(
            #     assignment__is_deleted=0,
            #     uploader_id__in=onco_A_user_ids,
            #     uploader__contact__in=contact_ids
            # ).exclude(
            #     uploader=193
            # ).annotate(
            #     custom_order=Case(
            #         When(uploader__contact__team__name='GSI', then=Value(1)),
            #         When(uploader__contact__team__name='CLUE', then=Value(2)),
            #         When(uploader__contact__team__name='etc', then=Value(4)),
            #         output_field=IntegerField()
            #     )
            # ).annotate(
            #     C1D1_count=Count('id', filter=gap_q3),
            #     gap1_count=Count('id', filter=gap_q1),
            #     gap2_count=Count('id', filter=gap_q2),
            #     gap1_total=Sum(
            #         ExpressionWrapper(
            #             Cast(F('create_date'), DateField()) - F('ICF_date'),
            #             output_field=IntegerField()
            #         ),
            #         filter=gap_q1
            #     ),
            #     gap2_total=Sum(
            #         ExpressionWrapper(
            #             Cast(F('create_date'), DateField()) - F('dosing_date'),
            #             output_field=IntegerField()
            #         ),
            #         filter=gap_q2
            #     )
            # ).annotate(
            #     total_gap=(Coalesce(F('gap1_total'), 0) + Coalesce(F('gap2_total'), 0)) /
            #               (Coalesce(F('gap1_count'), 0) + Coalesce(F('gap2_count'), 0) + Coalesce(F('C1D1_count'), 0)),
            #     visit_count=(Coalesce(F('gap1_count'), 0) + Coalesce(F('gap2_count'), 0) + Coalesce(F('C1D1_count'), 0))
            # ).values('uploader_id__first_name', 'uploader_id', 'total_gap', 'visit_count'
            #          ).order_by('custom_order', 'uploader_id__first_name')

            # ICF data input interval
            gap_condition_1 = Feedback.objects.filter(ICF_date__isnull=False, ICF_date__gte=from_month_date, ICF_date__lte=to_month_date)
            # cycle data input interval
            gap_condition_2 = Feedback.objects.filter(~Q(cycle='1', day='1') & ~Q(cycle='') & Q(ICF_date__isnull=True) &
                Q(dosing_date__gte=from_month_date) & Q(dosing_date__lte=to_month_date))
            # C1D1 data input interval
            gap_condition_3 = Feedback.objects.filter(Q(cycle='1', day='1') & Q(cycle__isnull=False) &
                Q(dosing_date__gte=from_month_date) & Q(dosing_date__lte=to_month_date))

            team_filter = Q(team__name=self.request.GET.get('team')) if self.request.GET.get('team') else Q()
            contact_ids = Contact.objects.filter(Q(onco_A=1) & team_filter).values_list('id', flat=True)

            gap_charts = Feedback.objects.filter(assignment__is_deleted=0,
                                                 uploader__contact__in=contact_ids) \
                    .values('uploader').distinct().order_by('uploader') \
                    .filter(~Q(uploader=193)) \
                    .annotate(custom_order=Case(When(uploader__contact__team__name='GSI', then=Value(1)),
                                                When(uploader__contact__team__name='CLUE', then=Value(2)),
                                                When(uploader__contact__team__name='etc', then=Value(4)),
                                                output_field=IntegerField())).order_by('custom_order',
                                                                                       'uploader_id__first_name') \
                    .annotate(C1D1_count=Count('id', filter=Q(id__in=gap_condition_3.values('id'))),
                              total_gap=(Coalesce(
                                  Sum(ExpressionWrapper(Cast(F('create_date'), DateField()) - F('ICF_date'),
                                                        output_field=IntegerField()) / 86400000000,
                                      filter=Q(id__in=gap_condition_1.values('id'))), 0) +
                                         Coalesce(Sum(
                                             ExpressionWrapper(Cast(F('create_date'), DateField()) - F('dosing_date'),
                                                               output_field=IntegerField()) / 86400000000,
                                             filter=Q(id__in=gap_condition_2.values('id'))), 0)) /
                                        (Count('id', filter=Q(id__in=gap_condition_1.values('id'))) +
                                         Count('id', filter=Q(id__in=gap_condition_2.values('id'))) + F('C1D1_count')),
                              visit_count=(Count('id', filter=Q(id__in=gap_condition_1.values('id'))) +
                                           Count('id', filter=Q(id__in=gap_condition_2.values('id'))) + F('C1D1_count'))
                              ).values('uploader_id__first_name', 'uploader_id', 'total_gap', 'visit_count')

            gap_crc_list = [];
            gap_count_list = [];
            visit_count_list = [];
            for gap in gap_charts:
                gap_crc_list.append({'id': str(gap['uploader_id']), 'name': str(gap['uploader_id__first_name'])})
                gap_count_list.append(str(gap['total_gap']))
                visit_count_list.append(str(gap['visit_count']))

            context['gap_crc_list'] = gap_crc_list
            context['gap_count_list'] = gap_count_list
            context['visit_count_list'] = visit_count_list

        # 첫 번째 폼
        if self.request.GET.get('from_date'):
            EOT_assign = Feedback.objects.filter(assignment__is_deleted=0, cycle='EOT').values('assignment_id')
            MAX_assign_status = STATUS_HISTORY.objects.filter(assignment__is_deleted=0) \
                .values('assignment_id') \
                .annotate(Max('create_date')).values('create_date__max')

            period_from_date = self.request.GET.get('from_date')
            period_to_date = self.request.GET.get('to_date')

            context['period_onco'] = ONCO_CR_COUNT.objects.filter(Q(research__is_deleted=False) & ~Q(research__status__in=['종료보고완료', '결과보고완료', '장기보관완료'])) \
                .annotate(custom_order=Case(When(research__is_recruiting='Recruiting', then=Value(1)),
                                            When(research__is_recruiting='Holding', then=Value(2)),
                                            When(research__is_recruiting='Not yet recruiting', then=Value(3)),
                                            When(research__is_recruiting='Completed', then=Value(4)),
                                            output_field=IntegerField()),
                          ATEAM=Case(When(research__onco_A='1', then=Value(1)),
                                     When(research__onco_A='0', then=Value(2)),
                                     output_field=IntegerField()),
                          pre_screening_period_filter=Count('research__assignment', distinct=True,
                                    filter=(Q(research__assignment__status='pre-screening', research__assignment__is_deleted=0) &
                                            Q(research__assignment__feedback__ICF_date__isnull=False) &
                                            Q(research__assignment__feedback__ICF_date__range=[period_from_date, period_to_date]))),
                          pre_screening_fail_period_filter=Count('research__assignment', distinct=True,
                                    filter=(Q(research__assignment__status='pre-screening-fail', research__assignment__is_deleted=0) &
                                            Q(research__assignment__feedback__scr_fail__isnull=False) &
                                            Q(research__assignment__feedback__scr_fail__range=[period_from_date, period_to_date]))),
                          screening_period_filter=Count('research__assignment', distinct=True,
                                    filter=(Q(research__assignment__feedback__ICF_date__range=[period_from_date, period_to_date]) &
                                            Q(research__assignment__phase=F('r_target'), research__assignment__is_deleted=0) &
                                            ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']))),
                          scr_fail_period_filter=Count('research__assignment', distinct=True,
                                    filter=(Q(research__assignment__feedback__scr_fail__range=[period_from_date, period_to_date]) &
                                            Q(research__assignment__phase=F('r_target'), research__assignment__is_deleted=0) &
                                            ~Q(research__assignment__status__in=['pre-screening', 'pre-screening-fail']))),
                          enroll_period_filter=Count('research__assignment', distinct=True,
                                    filter=(Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1',
                                              research__assignment__feedback__dosing_date__range=[period_from_date, period_to_date]) &
                                            Q(research__assignment__phase=F('r_target'), research__assignment__is_deleted=0))),
                          ongoing_period_filter=Count('research__assignment', distinct=True,
                                    filter=(((Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1',
                                               research__assignment__feedback__dosing_date__gte=period_from_date) &
                                              Q(research__assignment__feedback__cycle='EOT', research__assignment__feedback__dosing_date__lt=period_to_date)) |
                                             (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1',
                                               research__assignment__feedback__dosing_date__range=[period_from_date, period_to_date]) &
                                              ~Q(research__assignment__id__in=EOS_assign)) |  ##### 2024.03.07
                                             (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1',
                                               research__assignment__feedback__dosing_date__lte=period_from_date) &
                                             ~Q(research__assignment__id__in=EOT_assign) & ~Q(research__assignment__id__in=EOS_assign))) &  ##### 2024.03.07
                                              Q(research__assignment__phase=F('r_target'), research__assignment__is_deleted=0))),
                          FU_period_filter=Count('research__assignment', distinct=True,
                                    filter=Q(research__assignment__status_history__create_date__in=MAX_assign_status,
                                             research__assignment__status_history__create_date__range=[period_from_date, period_to_date],
                                             research__assignment__status_history__summary__contains='FU',
                                             research__assignment__phase=F('r_target'), research__assignment__is_deleted=0))) \
                .order_by('ATEAM', 'custom_order', 'research__research_name') \
                .prefetch_related('research', 'research__crc', 'research__cancer')

        today = date.today()
        context['today'] = today.strftime('%Y-%m-%d')
        context['first_day'] = datetime(today.year, today.month, 1).strftime('%Y-%m-%d')
        context['last_day'] = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1]).strftime('%Y-%m-%d')

        context['option_radio'] = self.request.GET.get('optionRadios')
        context['from_date'] = self.request.GET.get('from_date')
        context['to_date'] = self.request.GET.get('to_date')
        context['from_month_date'] = self.request.GET.get('from_month_date')
        context['to_month_date'] = self.request.GET.get('to_month_date')
        context['field_key_value'] = ONCO_CR_COUNT.create_field_value_and_text_dict()
        onco_A_crc = Contact.objects.filter(onco_A=1).values_list('user_id', flat=True)
        onco_A_investigator = InvestigatorContact.objects.filter(onco_A=1).values_list('user_id', flat=True)
        context['onco'] = onco_A_crc.union(onco_A_investigator)

        context['teams'] = Team.objects.all()
        context['team_name'] = team_name # 본인팀
        team_param = self.request.GET.get("team", None)
        context["selected_team"] = team_param if team_param is not None else team_name

        context['groupHeads'] = Contact.objects.filter(onco_A=1).filter(Q(user__groups__name='nurse') | Q(user__groups__name='medical records'))\
                                               .values('team')\
                                               .annotate(first_name=Min('name'))\
                                               .values('team__name', 'first_name')

        return context


@login_required()
def crc_ongoing_list(request, crc):
    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    
    ongo_condition1 = Feedback.objects.filter(cycle='1', day='1', dosing_date__lt=from_date).values('assignment_id')
    EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values('assignment_id')
    ongo_condition2 = Feedback.objects.filter(cycle='1', day='1', dosing_date__gte=from_date).values('assignment_id')
    EOT_assign2 = Feedback.objects.filter(Q(cycle='EOT', dosing_date__lt=to_date) & Q(assignment_id__in=ongo_condition2)).values('assignment_id')
    EOS_assign = Feedback.objects.exclude(eos__isnull=True).values('assignment_id').distinct()  ##### 2024.03.07

    crc_user = User.objects.filter(id=crc).values('id', 'first_name').first()
    crc_ongoing_condition = Feedback.objects.filter(assignment_id__is_deleted=0).filter(
            (Q(cycle='1', day='1', dosing_date__gte=from_date) & Q(cycle='EOT', dosing_date__lt=to_date)) |
            (Q(cycle='1', day='1', dosing_date__gte=from_date, dosing_date__lte=to_date) & ~Q(assignment_id__in=EOT_assign2) & ~Q(assignment_id__in=EOS_assign)) |  ##### 2024.03.07
            #(Q(cycle='EOT', dosing_date__year=date_year) & Q(cycle='EOT', dosing_date__month=date_month)) |
            (Q(cycle='1', day='1', dosing_date__lt=from_date) & ~Q(assignment_id__in=EOT_assign) & ~Q(assignment_id__in=EOS_assign))  ##### 2024.03.07
        ).values('assignment_id').distinct()
 
    crc_ongoing = Assignment.objects.filter(Q(curr_crc__user__id=crc_user['id']) & Q(id__in=crc_ongoing_condition) & ~Q(curr_crc__isnull=True)) \
                                    .annotate(type__count=Count('research__type', distinct=True),
                                              custom_order=Case(When(type__count=1, then=Value(1)),
                                                                When(type__count=2, then=Value(2)), output_field=IntegerField())) \
                                    .order_by('custom_order', 'research__research_name')\
                                    .prefetch_related('research__type')

    # EAP/PMS/완화연구/혈액/조직/기타 연구 1개 이상 가지고 있는 사람
    observe_and_PMS = Assignment.objects.filter(Q(id__in=crc_ongoing_condition) & (Q(research__type=3) | Q(research__type=4) | Q(research__type=5) | Q(research__type=6) | Q(research__type=7)))\
                                        .values('curr_crc').distinct().values_list('curr_crc__user__id', flat=True)

    return render(request, 'pages/research/crc_ongoing_list.html', {'crc_ongoing': crc_ongoing, 'crc_id': crc_user['id'], 'crc_name': crc_user['first_name'], 'observe_and_PMS': observe_and_PMS})


@login_required()
def crc_screening_list(request, crc):
    today = datetime.today()
    date_year = today.year
    date_month = today.month
    crc_user = User.objects.filter(id=crc).values('id', 'first_name').first()

    C1D1_assign = Feedback.objects.filter(Q(cycle='1', day='1', dosing_date__year=date_year, dosing_date__month=date_month)).values('assignment_id')
    crc_screening_condition = Feedback.objects.filter(assignment_id__is_deleted=0)\
                                              .filter((Q(ICF_date__year=date_year) & Q(ICF_date__month=date_month) &
                                                      ~Q(assignment__status='pre-screening') & ~Q(assignment__status='pre-screening-fail') &
                                                      ~Q(assignment_id__in=C1D1_assign))).values('assignment_id').distinct()

    crc_screening = Assignment.objects.filter(Q(curr_crc__user__id=crc_user['id']) & Q(id__in=crc_screening_condition) & ~Q(curr_crc__isnull=True)) \
                                      .annotate(type__count=Count('research__type', distinct=True),
                                                custom_order=Case(When(type__count=1, then=Value(1)),
                                                                  When(type__count=2, then=Value(2)), output_field=IntegerField())) \
                                      .order_by('custom_order', 'research__research_name')\
                                      .prefetch_related('research__type')

    # EAP/PMS/완화연구/혈액/조직/기타 연구 1개 이상 가지고 있는 사람
    observe_and_PMS = Assignment.objects.filter(Q(id__in=crc_screening_condition) & (Q(research__type=3) | Q(research__type=4) | Q(research__type=5) | Q(research__type=6) | Q(research__type=7)))\
                                        .values('curr_crc').distinct().values_list('curr_crc__user__id', flat=True)

    return render(request, 'pages/research/crc_screening_list.html', {'crc_screening': crc_screening, 'crc_id': crc_user['id'], 'crc_name': crc_user['first_name'], 'observe_and_PMS': observe_and_PMS})


@login_required()
def crc_input_gap_list(request, crc):
    # Input gap
    today = datetime.today()
    today_str = today.strftime('%Y-%m-%d')
    if not request.GET:
        from_date = datetime(today.year, today.month, 1)
        to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    elif request.GET.get('from_month_date') and request.GET.get('to_month_date'):
        from_date = request.GET.get('from_month_date')
        to_date = request.GET.get('to_month_date')

    crc_user = User.objects.filter(id=crc).values('id', 'first_name').first()
    onco_A = Contact.objects.filter(onco_A=1).values('id')
    gap_condition_1 = Feedback.objects.filter(ICF_date__isnull=False, ICF_date__gte=from_date, ICF_date__lte=to_date)

    gap_condition_2 = Feedback.objects.filter(~Q(cycle='1', day='1') & ~Q(cycle='') & Q(ICF_date__isnull=True) &
        Q(dosing_date__gte=from_date) & Q(dosing_date__lte=to_date))

    gap_condition_3 = Feedback.objects.filter(Q(cycle='1', day='1') & Q(cycle__isnull=False) &
        Q(dosing_date__gte=from_date) & Q(dosing_date__lte=to_date))

    input_gap_1 = Feedback.objects.filter(assignment__is_deleted=0, uploader__contact__in=onco_A)\
        .filter(Q(uploader_id=crc) & Q(id__in=gap_condition_1) & ~Q(assignment_id__curr_crc__isnull=True) |
                Q(uploader_id=crc) & Q(id__in=gap_condition_2) & ~Q(assignment_id__curr_crc__isnull=True)) \
        .annotate(total_gap=(Coalesce(Sum(ExpressionWrapper(Cast(F('create_date'), DateField()) - F('ICF_date'),
                    output_field=IntegerField()) / 86400000000, filter=Q(id__in=gap_condition_1.values('id'))), 0) +
                   Coalesce(Sum(ExpressionWrapper(Cast(F('create_date'), DateField()) - F('dosing_date'),
                    output_field=IntegerField()) / 86400000000, filter=Q(id__in=gap_condition_2.values('id'))), 0)))\
        .values('assignment_id__research_id__research_name', 'uploader_id__first_name',
                'assignment_id__name', 'assignment_id__register_number', 'assignment_id__research__id',
                'assignment_id__id', 'total_gap', 'create_date', 'dosing_date', 'ICF_date', 'cycle', 'day', 'assignment_id')\
        .order_by('-create_date')

    input_gap_2 = Feedback.objects.filter(assignment__is_deleted=0, uploader__contact__in=onco_A)\
        .filter(Q(uploader_id=crc) & Q(id__in=gap_condition_3) & ~Q(assignment_id__curr_crc__isnull=True)) \
        .annotate(C1D1_gap=(Coalesce(Sum(ExpressionWrapper(Cast(F('update_date'), DateField()) - F('dosing_date'),
                    output_field=IntegerField()) / 86400000000, filter=Q(id__in=gap_condition_3.values('id'))), 0)))\
        .values('assignment_id__research_id__research_name', 'uploader_id__first_name',
                'assignment_id__name', 'assignment_id__register_number', 'assignment_id__research__id',
                'assignment_id__id', 'C1D1_gap', 'update_date', 'dosing_date', 'ICF_date', 'cycle', 'day', 'assignment_id') \
        .order_by('-update_date')

    return render(request, 'pages/research/crc_input_gap_list.html',
                  {'input_gap_1': input_gap_1, 'input_gap_2': input_gap_2, 'crc_name': crc_user['first_name'], 'today': today_str})


@login_required()
def pi_research_list(request):
    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    date_year = today.year
    date_month = today.month
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    EOT_assign = Feedback.objects.filter(assignment__is_deleted=0, cycle='EOT').values('assignment_id')
    MAX_assign_status = STATUS_HISTORY.objects.filter(assignment__is_deleted=0) \
                                              .values('assignment_id') \
                                              .annotate(Max('create_date')).values('create_date__max')

    pi_research_list = ONCO_CR_COUNT.objects.filter(research__is_deleted=False, research__PI=request.user.first_name) \
        .annotate(custom_order=Case(When(research__is_recruiting='Recruiting', then=Value(1)), When(research__is_recruiting='Holding', then=Value(2)),
                                    When(research__is_recruiting='Not yet recruiting', then=Value(3)), When(research__is_recruiting='Completed', then=Value(4)), output_field=IntegerField()),
                  ATEAM=Case(When(research__onco_A='1', then=Value(1)), When(research__onco_A='0', then=Value(2)), output_field=IntegerField()),
                  screening_total_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__ICF_date__isnull=False) & Q(research__assignment__is_deleted=0) &
                                    ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail'))),
                  pre_screening_filter=Count('research__assignment', distinct=True,
                            filter=((Q(research__assignment__status='pre-screening') | Q(research__assignment__status='pre-screening-fail')) &
                                     Q(research__assignment__is_deleted=0) & Q(research__assignment__feedback__ICF_date__isnull=False))),
                  pre_screening_fail_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__status='pre-screening-fail') & Q(research__assignment__is_deleted=0) & Q(research__assignment__feedback__scr_fail__isnull=False))),
                  scr_fail_total_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__scr_fail__isnull=False) & Q(research__assignment__is_deleted=0) &
                                    ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail'))),
                  enroll_total_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__is_deleted=0))),
                  screening_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__ICF_date__isnull=False) & Q(research__assignment__phase=F('r_target'), research__assignment__curr_crc=F('research__crc')) &
                                    ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail') &
                                     Q(research__assignment__is_deleted=0) & Q(research__assignment__curr_crc=F('research__crc')))),
                  screening_month_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__phase=F('r_target'), research__assignment__curr_crc=F('research__crc'), research__assignment__is_deleted=0) &
                                    ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail') &
                                    ((Q(research__assignment__feedback__ICF_date__gte=from_date) &
                                      Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__lt=to_date)) |
                                    (Q(research__assignment__feedback__ICF_date__year=date_year) & Q(research__assignment__feedback__ICF_date__month=date_month))))),
                  scr_fail_month_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__scr_fail__gte=from_date) & Q(research__assignment__feedback__scr_fail__lte=to_date) &
                                    Q(research__assignment__phase=F('r_target'), research__assignment__curr_crc=F('research__crc'), research__assignment__is_deleted=0) &
                                    ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail'))),
                  scr_fail_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__scr_fail__isnull=False) &
                                    Q(research__assignment__phase=F('r_target'), research__assignment__curr_crc=F('research__crc'), research__assignment__is_deleted=0) &
                                    ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail'))),
                  ongoing_month_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__phase=F('r_target'), research__assignment__curr_crc=F('research__crc'), research__assignment__is_deleted=0) &
                                    ((Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__gte=from_date) &
                                      Q(research__assignment__feedback__cycle='EOT',  research__assignment__feedback__dosing_date__lt=to_date)) |
                                     (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__year=date_year) &
                                      Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__month=date_month))))),
                  ongoing_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__phase=F('r_target'), research__assignment__curr_crc=F('research__crc'), research__assignment__is_deleted=0) &
                                    ((Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__gte=from_date) &
                                      Q(research__assignment__feedback__cycle='EOT', research__assignment__feedback__dosing_date__lt=to_date)) |
                                    (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__year=date_year) &
                                     Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__month=date_month)) |
                                    (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__lt=from_date) &  ~Q(research__assignment__id__in=EOT_assign))))),
                  enroll_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1') &
                                    Q(research__assignment__phase=F('r_target'), research__assignment__curr_crc=F('research__crc'), research__assignment__is_deleted=0))),
                  FU_filter=Count('research__assignment', distinct=True,
                            filter=Q(research__assignment__status_history__create_date__in=MAX_assign_status, research__assignment__status_history__summary__contains='FU',
                                     research__assignment__phase=F('r_target'), research__assignment__is_deleted=0))
                  ) \
        .order_by('ATEAM', 'custom_order', 'research__research_name') \
        .values('research__id', 'research__research_name', 'research__is_recruiting', 'research__team', 'research__crc__name', 'id', 'r_target', 'research',
                'screening_total_filter', 'pre_screening_filter', 'pre_screening_fail_filter', 'scr_fail_total_filter', 'enroll_total_filter', 'FU_filter',
                'screening_filter', 'screening_month_filter', 'scr_fail_month_filter', 'scr_fail_filter', 'ongoing_month_filter', 'ongoing_filter', 'enroll_filter')

    pi_patient_list = Assignment.objects.filter(PI=request.user.first_name, is_deleted=0).order_by('research_id')\
                                        .values('research__id', 'research__research_name', 'id', 'no', 'status', 'register_number',
                                            'name', 'sex', 'age', 'dx', 'previous_tx', 'PI', 'create_date', 'update_date')
    pi_waiting_list = WaitingList.objects.filter(doctor=request.user.first_name, is_deleted=False).order_by('cancer_id')\
                                         .values('cancer', 'cancer__value', 'id', 'register_number', 'name', 'doctor', 'sex', 'age', 'curr_status', 'create_date', 'update_date')
    field_key_value = ONCO_CR_COUNT.create_field_value_and_text_dict()

    research_management = Research_Management.objects.filter(research__PI=request.user.first_name, is_deleted=0)
    research_management_field_key_value = Research_Management.create_field_value_and_text_dict()

    if request.GET.get('from_date'):
        period_from_date = request.GET.get('from_date')
        period_to_date = request.GET.get('to_date')
        EOT_assign = Feedback.objects.filter(assignment__is_deleted=0, cycle='EOT').values('assignment_id')
        C1D1_assign = Feedback.objects.filter(assignment__is_deleted=0, cycle='1', day='1').values('assignment_id')  # C1D1 존재
        scr_fail = Assignment.objects.filter(~Q(status='screening_fail', is_deleted=0))  # not screening fail
        MAX_assign_status = STATUS_HISTORY.objects.filter(assignment__is_deleted=0) \
                                                  .values('assignment_id') \
                                                  .annotate(Max('create_date')).values('create_date__max')

        pi_research_list_period = ONCO_CR_COUNT.objects.filter(research__is_deleted=False, research__PI=request.user.first_name) \
            .annotate(custom_order=Case(When(research__is_recruiting='Recruiting', then=Value(1)), When(research__is_recruiting='Holding', then=Value(2)),
                                        When(research__is_recruiting='Not yet recruiting', then=Value(3)), When(research__is_recruiting='Completed', then=Value(4)), output_field=IntegerField()),
                      ATEAM=Case(When(research__onco_A='1', then=Value(1)), When(research__onco_A='0', then=Value(2)), output_field=IntegerField()),
                      screening_period_filter=Count('research__assignment', distinct=True,
                            filter=(((Q(research__assignment__feedback__ICF_date__gte=period_from_date) &
                                      Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__lt=period_to_date)) |
                                     (Q(research__assignment__feedback__ICF_date__gte=period_from_date) & Q(research__assignment__feedback__ICF_date__lte=period_to_date)) |
                                     (Q(research__assignment__feedback__ICF_date__lte=period_from_date) &
                                      ~Q(research__assignment__id__in=C1D1_assign) & Q(research__assignment__id__in=scr_fail))) &
                                     (Q(research__assignment__phase=F('r_target'), research__assignment__is_deleted=0) &
                                      ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail')))),
                      scr_fail_period_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__scr_fail__gte=period_from_date) & Q(research__assignment__feedback__scr_fail__lte=period_to_date) &
                                    Q(research__assignment__phase=F('r_target'), research__assignment__is_deleted=0) &
                                    ~Q(research__assignment__status='pre-screening') & ~Q(research__assignment__status='pre-screening-fail'))),
                      enroll_period_filter=Count('research__assignment', distinct=True,
                            filter=(Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1',
                                      research__assignment__feedback__dosing_date__gte=period_from_date, research__assignment__feedback__dosing_date__lte=period_to_date) &
                                    Q(research__assignment__phase=F('r_target'),research__assignment__is_deleted=0))),
                      ongoing_period_filter=Count('research__assignment', distinct=True,
                            filter=(((Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__gte=period_from_date) &
                                      Q(research__assignment__feedback__cycle='EOT', research__assignment__feedback__dosing_date__lt=period_to_date)) |
                                     (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__gte=period_from_date) &
                                      Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__lte=period_to_date)) |
                                     (Q(research__assignment__feedback__cycle='EOT', research__assignment__feedback__dosing_date__gt=period_from_date) &
                                      Q(research__assignment__feedback__cycle='EOT', research__assignment__feedback__dosing_date__lte=period_to_date)) |
                                     (Q(research__assignment__feedback__cycle='1', research__assignment__feedback__day='1', research__assignment__feedback__dosing_date__lte=period_from_date) &
                                     ~Q(research__assignment__id__in=EOT_assign))) & Q(research__assignment__phase=F('r_target'), research__assignment__is_deleted=0))),
                      FU_period_filter=Count('research__assignment', distinct=True,
                            filter=Q(research__assignment__status_history__create_date__in=MAX_assign_status, research__assignment__status_history__create_date__gte=period_from_date, research__assignment__status_history__create_date__lte=period_to_date,
                                     research__assignment__status_history__summary__contains='FU', research__assignment__phase=F('r_target'),research__assignment__is_deleted=0))
                      ) \
            .order_by('ATEAM', 'custom_order', 'research__research_name') \
            .prefetch_related('research', 'research__crc')

        return render(request, 'pages/research/pi_research_list.html', {'pi_research_period': pi_research_list_period,
                                                                        'field_key_value': field_key_value,
                                                                        'option_radio': request.GET.get('optionRadios'),
                                                                        'from_date': request.GET.get('from_date'),
                                                                        'to_date': request.GET.get('to_date'),
                                                                        'pi_patient': pi_patient_list,
                                                                        'pi_waiting': pi_waiting_list})


    return render(request, 'pages/research/pi_research_list.html', {'pi_research': pi_research_list,
                                                                    'pi_patient': pi_patient_list,
                                                                    'pi_waiting': pi_waiting_list,
                                                                    'field_key_value': field_key_value,
                                                                    'option_radio': request.GET.get('optionRadios'),
                                                                    'from_date': request.GET.get('from_date'),
                                                                    'to_date': request.GET.get('to_date'),
                                                                    'research_management': research_management,
                                                                    'research_management_field_key_value': research_management_field_key_value})

def get_last_day_of_month(day):
    next_month = day.replace(day=28) + timedelta(days=4)
    return next_month - timedelta(days=next_month.day)

def chainer(s):
    return list(chain.from_iterable(s.str.split(' -> ')))

def compare_date(group, a, b):
    if len(group) > 2:
        col = [a] + group.iloc[2:len(group)]['create_date'].tolist() + [b]
    elif len(group) == 2:
        col = [a] + [b]
    else:
        col = np.nan
    group = group.assign(COMPARE_DATE=col)
    return group

@login_required()
def CRC_statistics(request):
    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    date_year = today.year
    date_month = today.month
    observe_and_PMS = Research.objects.filter(Q(type__value='EAP') | Q(type__value='PMS') | Q(type__value='Palliative') | Q(type__value='Blood') | Q(type__value='ETC')).values('id')

    DateSelectbox = request.GET.get('DateSelectbox')
    year_dropdown = [];
    for y in range(2021, (datetime.now().year + 1)):
        year_dropdown.append(str(y))

    is_recruiting_by_CRC = request.GET.get('is_recruiting_by_CRC')
    performance_year = request.GET.get('performance_year')
    performance_half_of_year = request.GET.get('performance_half_of_year')
    enroll_year = request.GET.get('enroll_year')
    enroll_year2 = request.GET.get('enroll_year2')

    # CRC별 Enroll 수 (월별)
    onco_A = Contact.objects.filter(onco_A=1).values('user_id')

    monthly_kwargs = {}
    technician_monthly_kwargs = {}
    for i in range(1, 13):
        gte = datetime(today.year, i, 1)
        lte = get_last_day_of_month(date(today.year, i, 1))
        mo = f'{gte:%b}'.lower()
        monthly_kwargs[mo] = Count('assignment_id', filter=(Q(dosing_date__gte=f'{gte:%Y-%m-%d}', dosing_date__lte=f'{lte:%Y-%m-%d}') & ~Q(assignment_id__research__id__in=observe_and_PMS)))
        technician_monthly_kwargs[mo] = Count('id', filter=Q(lab_date__gte=f'{gte:%Y-%m-%d}', lab_date__lte=f'{lte + timedelta(days=1):%Y-%m-%d}'))
    monthly_kwargs['total'] = Count('assignment_id', filter=(Q(dosing_date__year=today.year) & ~Q(assignment_id__research__id__in=observe_and_PMS)))
    technician_monthly_kwargs['total'] = Count('id', filter=Q(lab_date__year=today.year))
    for i in range(1, 13):
        gte = datetime(today.year, i, 1)
        lte = get_last_day_of_month(date(today.year, i, 1))
        mo = f'{gte:%b}'.lower()
        monthly_kwargs['PMS_' + mo] = Count('assignment_id', filter=Q(dosing_date__gte=f'{gte:%Y-%m-%d}', dosing_date__lte=f'{lte:%Y-%m-%d}', assignment_id__research__id__in=observe_and_PMS))
    monthly_kwargs['PMS_total'] = Count('assignment_id', filter=Q(dosing_date__year=today.year, assignment_id__research__id__in=observe_and_PMS))

    monthly_enroll = Feedback.objects.filter(Q(assignment__is_deleted=0, cycle='1', day='1', dosing_date__year=today.year)) \
                .values('uploader_id__first_name').distinct() \
                .order_by('uploader_id__first_name') \
                .annotate(**monthly_kwargs) \
                .values_list('uploader_id__first_name', *monthly_kwargs.keys(), 'uploader__contact__team__name', 'uploader__contact__onco_A') \
                .order_by('uploader_id__first_name')
    monthly_t_enroll = Supporting.objects.filter(Q(is_deleted=0, lab_date__year=today.year) & ~Q(technician='')) \
                .values('technician') \
                .order_by('technician') \
                .annotate(**technician_monthly_kwargs) \
                .annotate(technician_onco_A=Subquery(Contact.objects.filter(name=OuterRef('technician')).values('onco_A')[:1])) \
                .values_list('technician', *technician_monthly_kwargs.keys(), 'technician_onco_A')
    monthly_enroll_list = [list(i)[0:29] for i in monthly_enroll]
    monthly_t_enroll_list = [list(i)[0:15] for i in monthly_t_enroll]

    # CRC별 Cycle Visit 수 (월별)
    monthly_visit = Feedback.objects.filter(Q(assignment__is_deleted=0, dosing_date__year=today.year) & ~Q(cycle='')) \
        .values('uploader_id__first_name').distinct() \
        .order_by('uploader_id__first_name') \
        .annotate(**monthly_kwargs) \
        .values_list('uploader_id__first_name', *monthly_kwargs.keys(), 'uploader__contact__team__name', 'uploader__contact__onco_A') \
        .order_by('uploader_id__first_name')
    monthly_visit_list = [list(i)[0:29] for i in monthly_visit]

    # CRC별 담당 연구 개수 및 Annotation (ongoing condition)
    ongo_condition1 = Feedback.objects.filter(cycle='1', day='1').values('assignment_id')
    EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values('assignment_id')
    EOS_assign = Feedback.objects.exclude(eos__isnull=True).values('assignment_id').distinct() ##### 2024.03.07
    ongoing_condition = Feedback.objects.filter(
            (Q(cycle='1', day='1', dosing_date__gte=from_date) & Q(cycle='EOT', dosing_date__lt=to_date)) |
            (Q(cycle='1', day='1', dosing_date__year=date_year) & Q(cycle='1', day='1', dosing_date__month=date_month) & ~Q(assignment_id__in=EOS_assign)) | ##### 2024.03.07
            # (Q(cycle='EOT', dosing_date__year=date_year) & Q(cycle='EOT', dosing_date__month=date_month)) |
            (Q(cycle='1', day='1', dosing_date__lt=from_date) & ~Q(assignment_id__in=EOT_assign) & ~Q(assignment_id__in=EOS_assign))  ##### 2024.03.07
        ).values('assignment_id').distinct()

    ongoing = Research.objects.filter(is_deleted=0, crc__user_id__in=onco_A).values('crc__id').distinct() \
            .annotate(ongoing=Count('assignment', filter=(Q(assignment__in=ongoing_condition, assignment__is_deleted=0,
                                                            assignment__curr_crc__in=Contact.objects.filter(onco_A=1).values('id'),
                                                            assignment__curr_crc__id=F('crc__id'))), distinct=True)).order_by('crc__name')

    ongoing_list = [];
    for ongo in ongoing:
        ongoing_list.append(str(ongo['ongoing']))

    # 유형 수 (ex. IIT, PMS -> 2개 / IIT -> 1개)
    type_counts = Research.objects.filter(is_deleted=0)\
            .values('id') \
            .annotate(type_count=Count('type', distinct=True)).filter(type_count__gte=2) \
            .values('id')

    annotations = {}
    types = ('IIT', 'SIT', 'PMS', 'EAP', 'Palliative', 'Blood', 'ETC')
    for type in types:
        if type == 'IIT' or type == 'SIT':
            annotations[type] = Count('id', filter=Q(type__value=type), dinstinct=True)
            annotations[f'R_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Recruiting', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
            annotations[f'N_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Not yet recruiting', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
            annotations[f'H_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Holding', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
            annotations[f'C_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Completed', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
        else:
            annotations[type] = Count('id', filter=Q(type__value=type), dinstinct=True)
            annotations[f'R_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Recruiting', onco_A=1), dinstinct=True)
            annotations[f'N_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Not yet recruiting', onco_A=1), dinstinct=True)
            annotations[f'H_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Holding', onco_A=1), dinstinct=True)
            annotations[f'C_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Completed', onco_A=1), dinstinct=True)
        crc_research_counts = Research.objects.filter(is_deleted=0, crc__user_id__in=onco_A).values('crc')\
            .annotate(**annotations)\
            .values('crc__name', *annotations.keys()).order_by('crc__name')

    crc_research_count_dict = collections.defaultdict(list)
    for type_, count in itertools.product(types, crc_research_counts):
        crc_research_count_dict[type_].append(str(count[type_]))
        crc_research_count_dict[f'R_{type_}'].append(str(count[f'R_{type_}']))
        crc_research_count_dict[f'N_{type_}'].append(str(count[f'N_{type_}']))
        crc_research_count_dict[f'H_{type_}'].append(str(count[f'H_{type_}']))
        crc_research_count_dict[f'C_{type_}'].append(str(count[f'C_{type_}']))

    crc_list = [];  # CRC별 담당 연구 개수에서 crc__user_id__in=onco_A (Contact에서 onco_A=1)
    for count in crc_research_counts:
        crc_list.append(str(count['crc__name']))

    # CRC별/PI별 진행 환자 수
    N_of_ongoings_by_CRC_PI_annotations = {}
    investigators = InvestigatorContact.objects.filter(onco_A=True).values_list('name', flat=True)
    for investigator in investigators:
        N_of_ongoings_by_CRC_PI_annotations[investigator] = Count('assignment',
                filter=(Q(assignment__PI=investigator, assignment__in=ongoing_condition, assignment__is_deleted=0,
                          assignment__curr_crc__in=Contact.objects.filter(onco_A=1).values('id'), assignment__curr_crc__id=F('crc__id')) &
                       (Q(assignment__research__type__value='IIT') | Q(assignment__research__type__value='SIT')) & ~Q(assignment__research__in=type_counts)), distinct=True)

    N_of_ongoings_by_CRC_PI = Research.objects.filter(is_deleted=0, crc__user_id__in=onco_A, crc__user__groups__name='nurse')\
                                                  .values('crc') \
                                                  .annotate(**N_of_ongoings_by_CRC_PI_annotations) \
                                                  .values('crc__name', *N_of_ongoings_by_CRC_PI_annotations.keys()).order_by('crc__name')

    N_of_ongoings_by_CRC_PI_dict = collections.defaultdict(list)
    for investigator_, count in itertools.product(investigators, N_of_ongoings_by_CRC_PI):
        N_of_ongoings_by_CRC_PI_dict[investigator_].append(str(count[investigator_]))

    N_of_ongoings_by_CRC_PI_list = [];
    for crc in N_of_ongoings_by_CRC_PI:
        N_of_ongoings_by_CRC_PI_list.append(str(crc['crc__name']))

    # CRC별 Cycle Visit Count 증감율
    if today.month < 7:
        curr_from_date = datetime(today.year, 1, 1)
        curr_to_date = datetime(today.year, 6, 30)
        prev_from_date = datetime(today.year-1, 7, 1)
        prev_to_date = datetime(today.year-1, 12, 31)
    else:
        curr_from_date = datetime(today.year, 7, 1)
        curr_to_date = datetime(today.year, 12, 31)
        prev_from_date = datetime(today.year, 1, 1)
        prev_to_date = datetime(today.year, 6, 30)

    prev_variance = (Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A, dosing_date__gte=prev_from_date, dosing_date__lte=prev_to_date).count() +
                     Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A, ICF_date__gte=prev_from_date, ICF_date__lte=prev_to_date).count()) / float(181) / float(25)

    Contact.objects.filter(Q(onco_A=1) & ~Q(team__name='etc'))

    variance = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
        .values('uploader').distinct().order_by('uploader_id__first_name') \
        .annotate(curr_cycle=Count('id', filter=Q(dosing_date__gte=curr_from_date, dosing_date__lte=curr_to_date)),
                  curr_ICF=Count('id', filter=Q(ICF_date__gte=curr_from_date, ICF_date__lte=curr_to_date)),
                  curr=(F('curr_cycle') + F('curr_ICF')) / float((today - curr_from_date).days),
                  percentage=(F('curr') - prev_variance) / prev_variance * 100) \
        .values('uploader_id__first_name', 'percentage')

    xaxis_list = [];
    variance_list = [];
    for v in variance:
        xaxis_list.append(str(v['uploader_id__first_name']))
        variance_list.append(str(v['percentage']))

    # Performance
    if today.month < 7:
        from_date = datetime(today.year, 1, 1)
        to_date = datetime(today.year, 6, 30)
    else:
        from_date = datetime(today.year, 7, 1)
        to_date = datetime(today.year, 11, 30)

    ICF_count = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
        .values('assignment') \
        .annotate(ICF_count=Count('id', filter=Q(ICF_date__isnull=False)),
                  last_ICF_date=Max('ICF_date')) \
        .filter(ICF_count__gte=2) \
        .filter(last_ICF_date__range=[from_date, to_date]) \
        .values('assignment')

    performance = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
                    .values('uploader_id__first_name').distinct().order_by('uploader_id__first_name') \
                    .annotate(ICF=Count('assignment_id', filter=Q(ICF_date__isnull=False, ICF_date__gte=from_date, ICF_date__lte=to_date) &
                                                     ~Q(assignment__status='pre-screening') & ~Q(assignment__status='pre-screening-fail') &
                                                     ~Q(assignment__in=ICF_count) &
                                                     ~Q(assignment_id__research__id__in=observe_and_PMS), distinct=True) * Value(10),
                              pre_ICF=Count('id', filter=Q(ICF_date__isnull=False, ICF_date__gte=from_date, ICF_date__lte=to_date) &
                                                         ~Q(assignment__in=ICF_count) &
                                                         (Q(assignment__status='pre-screening') | Q(assignment__status='pre-screening-fail') |
                                                          Q(assignment_id__research__id__in=observe_and_PMS))) * Value(3),
                              re_ICF=Count('assignment', distinct=True, filter=Q(assignment__in=ICF_count)),
                              cycle_visit=Count('id', filter=Q(dosing_date__gte=from_date, dosing_date__lte=to_date) & ~Q(cycle=''))) \
                    .values_list('uploader_id__first_name', 'ICF', 'pre_ICF', 're_ICF', 'cycle_visit', 'uploader__groups__name', 'uploader_id', 'uploader__contact__team__name')
    performance_list = [list(i)[0:8] for i in performance]

    # setup 수
    workers = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=from_date, initiation_date__lte=to_date, set_up__isnull=False) \
        .values('id') \
        .annotate(worker=Count('set_up')) \
        .values('id', 'worker')

    # 유형 수 (ex. IIT, PMS -> 2개 / IIT -> 1개) 2개 이상인 연구
    type_counts = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=from_date, initiation_date__lte=to_date, set_up__isnull=False) \
                            .values('id') \
                            .annotate(type_count=Count('type', distinct=True)).filter(type_count__gte=2) \
                            .values('id')

    performance_setup = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=from_date, initiation_date__lte=to_date, set_up__isnull=False) \
        .values('set_up') \
        .annotate(IIT_setup=Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=1).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(120), IntegerField()) +
                            Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=2).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(60), IntegerField()) +
                            Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=3).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(40), IntegerField()),
                  SIT_setup=Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=1).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(30), IntegerField()) +
                            Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=2).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(15), IntegerField()) +
                            Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=3).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(10), IntegerField()),
                  ETC_setup=Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                    (Q(id__in=workers.filter(worker=1).values('id'))), distinct=True) * Value(10), IntegerField()) +
                            Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                    (Q(id__in=workers.filter(worker=2).values('id'))), distinct=True) * Value(5), IntegerField()) +
                            Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                    (Q(id__in=workers.filter(worker=3).values('id'))), distinct=True) * Value(3.3), IntegerField())) \
        .values('set_up__user_id', 'IIT_setup', 'SIT_setup', 'ETC_setup')

    feedback_uploader = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
                               .values('uploader_id').distinct().values_list('uploader_id', flat=True)
    subject_of_performance = Contact.objects.filter(Q(onco_A=1) & (Q(user_id__groups__name='nurse') | Q(user_id__groups__name='medical records') | Q(user_id__groups__name='SETUP')))\
                                            .values_list('user__id', flat=True)

    s = set(feedback_uploader)
    none_feedback_uploader = [x for x in subject_of_performance if x not in s]
    for n_f_u in none_feedback_uploader:
        name = User.objects.filter(id=n_f_u).values_list('first_name', 'groups__name', 'id')
        performance_list.append([name[0][0], 0, 0, 0, 0, name[0][1], name[0][2]])
        performance_list.sort()

    setting_new = {dct['set_up__user_id']: itemgetter('IIT_setup', 'SIT_setup', 'ETC_setup')(dct) for dct in performance_setup}
    uploader_ids = map(lambda x: x[-2], performance_list)
    performance_setup_list = [[i, *setting_new.get(i, (0, 0, 0))] for i in uploader_ids]

    performance_technician = Supporting.objects.filter(Q(is_deleted=0, lab_date__range=(from_date, to_date + timedelta(days=1))) & ~Q(technician='')) \
        .values('technician') \
        .annotate(PK_Sampling=Count('id', filter=Q(lab_date__gte=from_date, lab_date__lte=to_date + timedelta(days=1)))) \
        .values_list('technician', 'PK_Sampling')
    performance_technician_list = [list(i)[0:2] for i in performance_technician]

    tab = ''

    if 'enroll_year' in request.GET:  # CRC별 Enroll 수 (월별) 필터링
        enroll_year = request.GET.get('enroll_year')
        tab = 'status04_02'

        monthly_kwargs = {}
        technician_monthly_kwargs = {}
        for i in range(1, 13):
            gte = datetime(int(enroll_year), i, 1)
            lte = get_last_day_of_month(date(int(enroll_year), i, 1))
            mo = f'{gte:%b}'.lower()
            monthly_kwargs[mo] = Count('assignment_id', filter=(Q(dosing_date__gte=f'{gte:%Y-%m-%d}', dosing_date__lte=f'{lte:%Y-%m-%d}') & ~Q(assignment_id__research__id__in=observe_and_PMS)))
            technician_monthly_kwargs[mo] = Count('id', filter=Q(lab_date__gte=f'{gte:%Y-%m-%d}', lab_date__lte=f'{lte + timedelta(days=1):%Y-%m-%d}'))
        monthly_kwargs['total'] = Count('assignment_id', filter=(Q(dosing_date__year=enroll_year) & ~Q(assignment_id__research__id__in=observe_and_PMS)))
        technician_monthly_kwargs['total'] = Count('id', filter=Q(lab_date__year=enroll_year))
        for i in range(1, 13):
            gte = datetime(int(enroll_year), i, 1)
            lte = get_last_day_of_month(date(int(enroll_year), i, 1))
            mo = f'{gte:%b}'.lower()
            monthly_kwargs['PMS_' + mo] = Count('assignment_id', filter=Q(dosing_date__gte=f'{gte:%Y-%m-%d}', dosing_date__lte=f'{lte:%Y-%m-%d}', assignment_id__research__id__in=observe_and_PMS))
        monthly_kwargs['PMS_total'] = Count('assignment_id', filter=Q(dosing_date__year=enroll_year, assignment_id__research__id__in=observe_and_PMS))

        monthly_enroll = Feedback.objects.filter(Q(assignment__is_deleted=0, cycle='1', day='1', dosing_date__year=enroll_year)) \
            .values('uploader_id__first_name').distinct() \
            .order_by('uploader_id__first_name') \
            .annotate(**monthly_kwargs) \
            .values_list('uploader_id__first_name', *monthly_kwargs.keys(), 'uploader__contact__team__name', 'uploader__contact__onco_A') \
            .order_by('uploader_id__first_name')
        monthly_t_enroll = Supporting.objects.filter(Q(is_deleted=0, lab_date__year=enroll_year) & ~Q(technician='')) \
            .values('technician') \
            .order_by('technician') \
            .annotate(**technician_monthly_kwargs) \
            .annotate(technician_onco_A=Subquery(Contact.objects.filter(name=OuterRef('technician')).values('onco_A')[:1])) \
            .values_list('technician', *technician_monthly_kwargs.keys(), 'technician_onco_A')
        monthly_enroll_list = [list(i)[0:29] for i in monthly_enroll]
        monthly_t_enroll_list = [list(i)[0:15] for i in monthly_t_enroll]

    if 'enroll_year2' in request.GET:  # CRC별 Cycle Visit 수 필터링
        enroll_year2 = request.GET.get('enroll_year2')
        tab = 'status04_03'

        monthly_kwargs = {}
        technician_monthly_kwargs = {}
        for i in range(1, 13):
            gte = datetime(int(enroll_year2), i, 1)
            lte = get_last_day_of_month(date(int(enroll_year2), i, 1))
            mo = f'{gte:%b}'.lower()
            monthly_kwargs[mo] = Count('assignment_id', filter=(Q(dosing_date__gte=f'{gte:%Y-%m-%d}', dosing_date__lte=f'{lte:%Y-%m-%d}') &  ~Q(assignment_id__research__id__in=observe_and_PMS)))
            technician_monthly_kwargs[mo] = Count('id', filter=Q(lab_date__gte=f'{gte:%Y-%m-%d}', lab_date__lte=f'{lte + timedelta(days=1):%Y-%m-%d}'))
        monthly_kwargs['total'] = Count('assignment_id', filter=(Q(dosing_date__year=enroll_year2) & ~Q(assignment_id__research__id__in=observe_and_PMS)))
        technician_monthly_kwargs['total'] = Count('id', filter=Q(lab_date__year=enroll_year2))
        for i in range(1, 13):
            gte = datetime(int(enroll_year2), i, 1)
            lte = get_last_day_of_month(date(int(enroll_year2), i, 1))
            mo = f'{gte:%b}'.lower()
            monthly_kwargs['PMS_' + mo] = Count('assignment_id', filter=Q(dosing_date__gte=f'{gte:%Y-%m-%d}', dosing_date__lte=f'{lte:%Y-%m-%d}', assignment_id__research__id__in=observe_and_PMS))
        monthly_kwargs['PMS_total'] = Count('assignment_id', filter=Q(dosing_date__year=enroll_year2, assignment_id__research__id__in=observe_and_PMS))

        monthly_visit = Feedback.objects.filter(Q(assignment__is_deleted=0, dosing_date__year=enroll_year2) & ~Q(cycle='')) \
            .values('uploader_id__first_name').distinct() \
            .order_by('uploader_id__first_name') \
            .annotate(**monthly_kwargs) \
            .values_list('uploader_id__first_name', *monthly_kwargs.keys(), 'uploader__contact__team__name', 'uploader__contact__onco_A') \
            .order_by('uploader_id__first_name')
        monthly_visit_list = [list(i)[0:29] for i in monthly_visit]

    if 'from_date' in request.GET:
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        tab = 'status04_01'
        oncoA = Research.objects.filter(Q(is_deleted=0, crc__user_id__in=onco_A) &
                                             (Q(crc__user_id__groups__name='nurse') | Q(crc__user_id__groups__name='medical records') | Q(crc__user_id__groups__name='SETUP'))) \
                             .annotate(custom_order=Case(When(crc__team__name='GSI', then=Value(1)), When(crc__team__name='CLUE', then=Value(2)),
                                                         When(crc__team__name='etc', then=Value(3)), output_field=IntegerField()))\
                             .values('crc__name').distinct().order_by('custom_order', 'crc__name')  # CRC, 의무기록사, SETUP
        oncoA_list = [];
        for onco in oncoA:
            oncoA_list.append(str(onco['crc__name'])) # x axis
        crc_list = oncoA_list

        latest_dates = History.objects.filter(research__is_deleted=0).values('research_id') \
            .annotate(latest_created=Max('create_date', filter=(Q(summary__field_summary__crc__isnull=False, create_date__lte=from_date))))
        histories = History.objects.filter(research__is_deleted=0).values('research_id') \
            .filter(Q(summary__field_summary__crc__isnull=False, create_date__gte=from_date, create_date__lte=to_date) |
                    Q(create_date__in=latest_dates.values('latest_created'))) \
            .annotate(create_date_str=Cast(F('create_date'), DateField())) \
            .values('research__research_name', 'summary__field_summary__crc', 'create_date_str').order_by('research__research_name')

        if not histories.exists():
            count = ['0'] * len(oncoA_list)
        else:
            df = pd.DataFrame.from_records(histories).sort_values('research__research_name').reset_index(drop=True)
            research_count = df['research__research_name'].value_counts()
            unique = research_count[research_count == 1].index.values
            df['create_date'] = df['create_date_str'].apply(pd.to_datetime, utc=True)
            df = df.loc[(df['create_date'] >= from_date) & (df['create_date'] <= to_date) | (df['research__research_name'].isin(unique))].reset_index(drop=True)
            # df = df.loc[df.groupby(['research__research_name'])['create_date'].idxmax()].reset_index(drop=True)
            df['summary__field_summary__crc'] = df['summary__field_summary__crc'].str.replace('[', '').str.replace(']', '')

            lengths = df['summary__field_summary__crc'].str.split(' -> ').map(len)
            df = pd.DataFrame({'research_name': np.repeat(df['research__research_name'], lengths),
                               'CRC': chainer(df['summary__field_summary__crc']),
                               'create_date': np.repeat(df['create_date'], lengths)}).reset_index(drop=True)

            period = df['create_date'].between(from_date, to_date)
            df2 = df.loc[df.assign(tmp=df.index)[~period].groupby('research_name')['tmp'].idxmax()]
            df = pd.concat([df[period], df2])

            # df = df[~df['CRC'].str.contains('None', na=False, case=False)]
            df = df[~(df['create_date'] == from_date)].reset_index(drop=True).drop_duplicates(subset=['research_name', 'CRC'])

            count = df.groupby('research_name').apply(compare_date, a=from_date, b=to_date).reset_index(drop=True)
            count['COMPARE_DATE'] = count['COMPARE_DATE'].apply(pd.to_datetime, utc=True)
            count['INTERVAL_DAYS'] = np.where(count['create_date'].between(from_date, to_date),
                        abs(count['COMPARE_DATE'].sub(count['create_date'])).dt.days.div((datetime.strptime(to_date, '%Y-%m-%d') - datetime.strptime(from_date, '%Y-%m-%d')).days),
                        int(1))

            count['CRC'] = count['CRC'].str.strip('[]')
            count = count.assign(CRC=count['CRC'].str.split('/|, ')).explode('CRC')
            count = count[count['CRC'].isin(oncoA_list)]
            count['INTERVAL_DAYS'] /= count.groupby(level=0)['INTERVAL_DAYS'].transform('size')
            count = round(count.groupby('CRC')['INTERVAL_DAYS'].sum(), 2).reset_index(name='WORKING_COUNTS')
            count = count.set_index('CRC').reindex(index=oncoA_list).fillna(0).reset_index()
            count = count.WORKING_COUNTS.tolist()

    if 'half_of_year' in request.GET:
        performance_year = request.GET.get('performance_year')
        performance_half_of_year = request.GET.get('half_of_year')
        tab = 'status04_06'
        if performance_half_of_year == 'first_half':
            performance_from_date = datetime(int(performance_year), 1, 1)
            performance_to_date = datetime(int(performance_year), 6, 30)
        elif performance_half_of_year == 'second_half':
            performance_from_date = datetime(int(performance_year), 7, 1)
            performance_to_date = datetime(int(performance_year), 12, 31)

        ICF_count = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
            .values('assignment') \
            .annotate(ICF_count=Count('id', filter=Q(ICF_date__isnull=False)),
                      last_ICF_date=Max('ICF_date')) \
            .filter(ICF_count__gte=2) \
            .filter(last_ICF_date__range=[performance_from_date, performance_to_date]) \
            .values('assignment')

        performance = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
            .values('uploader_id__first_name').distinct().order_by('uploader_id__first_name') \
            .annotate(ICF=Count('assignment_id', filter=Q(ICF_date__isnull=False, ICF_date__gte=performance_from_date, ICF_date__lte=performance_to_date) &
                                             ~Q(assignment__status='pre-screening') & ~Q(assignment__status='pre-screening-fail') &
                                             ~Q(assignment__in=ICF_count) &
                                             ~Q(assignment_id__research__id__in=observe_and_PMS), distinct=True) * Value(10),
                      pre_ICF=Count('id', filter=Q(ICF_date__isnull=False, ICF_date__gte=performance_from_date, ICF_date__lte=performance_to_date) &
                                                 ~Q(assignment__in=ICF_count) &
                                                 (Q(assignment__status='pre-screening') | Q(assignment__status='pre-screening-fail') | Q(assignment_id__research__id__in=observe_and_PMS))) * Value(3),
                      re_ICF=Count('assignment', distinct=True, filter=Q(assignment__in=ICF_count)),
                      cycle_visit=Count('id', filter=Q(dosing_date__gte=performance_from_date, dosing_date__lte=performance_to_date) & ~Q(cycle=''))) \
            .values_list('uploader_id__first_name', 'ICF', 'pre_ICF', 're_ICF', 'cycle_visit', 'uploader__groups__name', 'uploader_id', 'uploader__contact__team__name')
        performance_list = [list(i)[0:8] for i in performance]

        workers = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=performance_from_date, initiation_date__lte=performance_to_date, set_up__isnull=False) \
            .values('id') \
            .annotate(worker=Count('set_up')) \
            .values('id', 'worker')

        type_counts = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=performance_from_date, initiation_date__lte=performance_to_date, set_up__isnull=False) \
            .values('id') \
            .annotate(type_count=Count('type', distinct=True)).filter(type_count__gte=2) \
            .values('id')

        performance_setup = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=performance_from_date, initiation_date__lte=performance_to_date, set_up__isnull=False) \
            .values('set_up') \
            .annotate(IIT_setup=Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=1).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(120), IntegerField()) +
                                Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=2).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(60), IntegerField()) +
                                Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=3).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(40), IntegerField()),
                      SIT_setup=Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=1).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(30), IntegerField()) +
                                Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=2).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(15), IntegerField()) +
                                Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=3).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(10), IntegerField()),
                      ETC_setup=Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                        (Q(id__in=workers.filter(worker=1).values('id'))), distinct=True) * Value(10), IntegerField()) +
                                Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                        (Q(id__in=workers.filter(worker=2).values('id'))), distinct=True) * Value(5), IntegerField()) +
                                Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                        (Q(id__in=workers.filter(worker=3).values('id'))), distinct=True) * Value(3.3), IntegerField())) \
            .values('set_up__user_id', 'IIT_setup', 'SIT_setup', 'ETC_setup')

        feedback_uploader = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
                                            .values('uploader_id').distinct().values_list('uploader_id', flat=True)
        subject_of_performance = Contact.objects.filter(Q(onco_A=1) & (Q(user_id__groups__name='nurse') | Q(user_id__groups__name='medical records') | Q(user_id__groups__name='SETUP'))) \
                                                .values_list('user__id', flat=True)

        s = set(feedback_uploader)
        none_feedback_uploader = [x for x in subject_of_performance if x not in s]
        for n_f_u in none_feedback_uploader:
            name = User.objects.filter(id=n_f_u).values_list('first_name', 'groups__name', 'id')
            performance_list.append([name[0][0], 0, 0, 0, 0, name[0][1], name[0][2]])
            performance_list.sort()

        setting_new = {dct['set_up__user_id']: itemgetter('IIT_setup', 'SIT_setup', 'ETC_setup')(dct) for dct in performance_setup}
        uploader_ids = map(lambda x: x[-2], performance_list)
        performance_setup_list = [[i, *setting_new.get(i, (0, 0, 0))] for i in uploader_ids]

        performance_technician = Supporting.objects.filter(Q(is_deleted=0, lab_date__range=(performance_from_date, performance_to_date + timedelta(days=1))) & ~Q(technician='')) \
                                                   .values('technician') \
                                                   .annotate(PK_Sampling=Count('id', filter=Q(lab_date__gte=performance_from_date, lab_date__lte=performance_to_date + timedelta(days=1)))) \
                                                   .values_list('technician', 'PK_Sampling')
        performance_technician_list = [list(i)[0:2] for i in performance_technician]

        #onco_A_crc = Research.objects.filter(is_deleted=0, crc__user_id__in=onco_A)\
        #                             .annotate(custom_order=Case(When(crc__team__name='GSI', then=Value(1)), When(crc__team__name='CLUE', then=Value(2)),
        #                                                         When(crc__team__name='etc', then=Value(3)), output_field=IntegerField()))\
        #                             .values('crc__name').distinct().order_by('custom_order', 'crc__name')

        #crc_list = [];
        #for crc in onco_A_crc:
        #    crc_list.append(str(crc['crc__name']))

    return render(request, 'pages/research/statistics/CRC_statistics.html',
                  {
                      'today': today,
                      'tab': tab,
                      'year_dropdown': year_dropdown,
                      'date_year': today.strftime('%Y'),
                      'teams': Team.objects.all(),
                      # CRC별 진행 환자 수
                      'N_of_ongoings_by_CRC_PI_dict': N_of_ongoings_by_CRC_PI_dict.items(),
                      'N_of_ongoings_by_CRC_PI_list': N_of_ongoings_by_CRC_PI_list,
                      'count': count,
                      # CRC별 담당 연구 개수
                      'is_recruiting_by_CRC': is_recruiting_by_CRC,
                      'crc_research_count_dict': crc_research_count_dict,
                      'ongoing': ongoing_list,
                      'crc': crc_list,
                      'DateSelectbox': DateSelectbox,
                      'from_date': from_date,
                      'to_date': to_date,
                      # CRC별 Enroll 수 (월별)
                      'monthly_enroll_list': monthly_enroll_list,
                      'monthly_t_enroll_list': monthly_t_enroll_list,
                      'enroll_year': enroll_year if enroll_year else today.year,
                      # CRC별 Cycle Visit 수 (월별)
                      'monthly_visit_list': monthly_visit_list,
                      'enroll_year2': enroll_year2 if enroll_year2 else today.year,
                      # CRC별 Cycle Visit Count 증감율
                      'variance_xaxis_list': xaxis_list,
                      'variance_list': variance_list,
                      'prev_variance': prev_variance,
                      # Performance
                      'performance_year': performance_year,
                      'performance_half_of_year': performance_half_of_year,
                      'performance_list': performance_list,
                      'performance_t_list': performance_technician_list,
                      'performance_setup_list': performance_setup_list,
                  })

@login_required()
def PI_statistics(request):
    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    date_year = today.year
    date_month = today.month
    investigators = InvestigatorContact.objects.filter(onco_A=True).values_list('name', flat=True)

    # ETC 판단용: 2개 이상 type 가진 연구들
    type_counts = Research.objects.filter(is_deleted=0) \
        .values('id') \
        .annotate(type_count=Count('type', distinct=True)) \
        .filter(type_count__gte=2) \
        .values_list('id', flat=True)

    # PI별/CRC별 진행 환자 수 (그래프, 표)
    ongo_condition1 = Feedback.objects.filter(cycle='1', day='1').values('assignment_id')
    EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values('assignment_id')

    # Feedback 기준 진행 환자 assignment id
    ongoing_assignment_ids = Feedback.objects.filter(
         (Q(cycle='1', day='1', dosing_date__gte=from_date) & Q(cycle='EOT', dosing_date__lt=to_date)) |
         (Q(cycle='1', day='1', dosing_date__year=date_year) & Q(cycle='1', day='1', dosing_date__month=date_month)) |
         (Q(cycle='1', day='1', dosing_date__lt=from_date) & ~Q(assignment_id__in=EOT_assign))
    ).values_list('assignment_id', flat=True).distinct()

    # 대상 CRC 목록
    coordinators = Contact.objects.filter(
        onco_A=True,
        user_id__groups__name__in=['nurse', 'medical records']
    ).values('id', 'name', 'team__name').distinct().order_by('name')

    # 연구 타입별 ID 추출
    iit_ids = set(Research.objects.filter(is_deleted=0, type__value='IIT').values_list('id', flat=True))
    sit_ids = set(Research.objects.filter(is_deleted=0, type__value='SIT').values_list('id', flat=True))
    etc_ids = set(type_counts)

    # 데이터 저장용
    N_of_ongoings_by_PI_CRC_dict = collections.defaultdict(lambda: collections.defaultdict(lambda: {'IIT': 0, 'SIT': 0, 'ETC': 0}))
    pi_set = set()

    # 데이터 구성
    for c in coordinators:
        crc_id = c['id']
        crc_name = c['name']
        team_name = c['team__name']
        key = (crc_name, team_name)

        assignments = Assignment.objects.filter(curr_crc=crc_id, is_deleted=False, id__in=ongoing_assignment_ids, research__PI__in=investigators)

        for a in assignments:
            research = a.research
            if not research or research.is_deleted:
                continue

            pi = research.PI or "(미입력)"
            pi_set.add(pi)

            if research.id in etc_ids:
                N_of_ongoings_by_PI_CRC_dict[key][pi]['ETC'] += 1
            elif research.id in iit_ids:
                N_of_ongoings_by_PI_CRC_dict[key][pi]['IIT'] += 1
            elif research.id in sit_ids:
                N_of_ongoings_by_PI_CRC_dict[key][pi]['SIT'] += 1

    # PI 정렬
    N_of_ongoings_by_PI_CRC_list = sorted(pi_set)

    # 누락된 PI 보정
    for key in N_of_ongoings_by_PI_CRC_dict:
        for pi in N_of_ongoings_by_PI_CRC_list:
            if pi not in N_of_ongoings_by_PI_CRC_dict[key]:
                N_of_ongoings_by_PI_CRC_dict[key][pi] = {'IIT': 0, 'SIT': 0, 'ETC': 0}

    # 그래프용 series (CRC 기준 시리즈 생성)
    series_js = []
    for (crc_name, team_name), pi_data in N_of_ongoings_by_PI_CRC_dict.items():
        for type_label in ['IIT', 'SIT', 'ETC']:
            data_array = [str(pi_data[pi][type_label]) for pi in N_of_ongoings_by_PI_CRC_list]
            series_js.append(
                f"{{name: \"{crc_name} ({type_label})\", group: \"{type_label}\", data: {data_array}}}"
            )
    series_js_string = "[" + ",".join(series_js) + "]"

    # table용 데이터: dict[(crc, team)] → list of 3*N PI counts (IIT, SIT, ETC 순)
    N_of_ongoings_by_PI_CRC_table_dict = {}

    for key, pi_dict in N_of_ongoings_by_PI_CRC_dict.items():
        row_by_type = {
            'IIT': [],
            'SIT': [],
            'ETC': [],
        }
        for type_label in ['IIT', 'SIT', 'ETC']:
            for pi in N_of_ongoings_by_PI_CRC_list:
                row_by_type[type_label].append(pi_dict[pi][type_label])
        N_of_ongoings_by_PI_CRC_table_dict[key] = row_by_type

    # (25/08/13 추가) 타입별 PI 합계
    totals_by_type = {'IIT': [0] * len(N_of_ongoings_by_PI_CRC_list),
                      'SIT': [0] * len(N_of_ongoings_by_PI_CRC_list),
                      'ETC': [0] * len(N_of_ongoings_by_PI_CRC_list)}

    for row in N_of_ongoings_by_PI_CRC_table_dict.values():
        for t in ['IIT', 'SIT', 'ETC']:
            for i, v in enumerate(row[t]):
                totals_by_type[t][i] += v

    # PI별 담당 연구 개수
    PI_annotations = {}
    types = ('IIT', 'SIT', 'PMS', 'EAP', 'Palliative', 'Blood', 'ETC')
    for type in types:
        if type == 'IIT' or type == 'SIT':
            PI_annotations[type] = Count('id', filter=(Q(type__value=type, onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
            PI_annotations[f'R_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Recruiting', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
            PI_annotations[f'N_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Not yet recruiting', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
            PI_annotations[f'H_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Holding', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
            PI_annotations[f'C_{type}'] = Count('id', filter=(Q(type__value=type, is_recruiting='Completed', onco_A=1) & ~Q(id__in=type_counts)), dinstinct=True)
        else:
            PI_annotations[type] = Count('id', filter=(Q(type__value=type, onco_A=1)), dinstinct=True)
            PI_annotations[f'R_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Recruiting', onco_A=1), dinstinct=True)
            PI_annotations[f'N_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Not yet recruiting', onco_A=1), dinstinct=True)
            PI_annotations[f'H_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Holding', onco_A=1), dinstinct=True)
            PI_annotations[f'C_{type}'] = Count('id', filter=Q(type__value=type, is_recruiting='Completed', onco_A=1), dinstinct=True)

        PI_research_counts = Research.objects.filter(is_deleted=0, PI__in=investigators).values('PI') \
            .annotate(**PI_annotations) \
            .values('PI', *PI_annotations.keys()).order_by('PI')

    PI_research_count_dict = collections.defaultdict(list)
    for type_, count in itertools.product(types, PI_research_counts):
        PI_research_count_dict[type_].append(str(count[type_]))
        PI_research_count_dict[f'R_{type_}'].append(str(count[f'R_{type_}']))
        PI_research_count_dict[f'N_{type_}'].append(str(count[f'N_{type_}']))
        PI_research_count_dict[f'H_{type_}'].append(str(count[f'H_{type_}']))
        PI_research_count_dict[f'C_{type_}'].append(str(count[f'C_{type_}']))

    PI_list = [];
    for count in PI_research_counts:
        PI_list.append(str(count['PI']))

    # 교수별 Enroll 수 (월별)
    monthly_by_PI_kwargs = {}
    for i in range(1, 13):
        gte = datetime(today.year, i, 1)
        lte = get_last_day_of_month(date(today.year, i, 1))
        mo = f'{gte:%b}'.lower()
        monthly_by_PI_kwargs[mo] = Count('assignment_id', filter=(Q(dosing_date__gte=f'{gte:%Y-%m-%d}', dosing_date__lte=f'{lte:%Y-%m-%d}')))
    monthly_by_PI_kwargs['total'] = Count('assignment_id', filter=(Q(dosing_date__year=today.year)))

    monthly_enroll_by_PI = Feedback.objects.filter(Q(assignment__is_deleted=0, cycle='1', day='1', dosing_date__year=today.year, assignment__PI__in=investigators)) \
            .values('assignment__PI').distinct() \
            .order_by('assignment__PI') \
            .annotate(**monthly_by_PI_kwargs) \
            .values_list('assignment__PI', *monthly_by_PI_kwargs.keys()) \
            .order_by('assignment__PI')
    monthly_enroll_by_PI_list = [list(i)[0:28] for i in monthly_enroll_by_PI]

    return render(request, 'pages/research/statistics/PI_statistics.html',
                  {
                      'today': today,
                      'PI_research_count_dict': PI_research_count_dict,
                      "series_js": series_js_string,
                      "pi_list": json.dumps(N_of_ongoings_by_PI_CRC_list, ensure_ascii=False),  # PI 리스트
                      "N_of_ongoings_by_PI_CRC_list": N_of_ongoings_by_PI_CRC_list,
                      'N_of_ongoings_by_PI_CRC_table_dict': N_of_ongoings_by_PI_CRC_table_dict,
                      'totals_by_type': totals_by_type,
                      "type_labels": [
                            ("IIT", "IIT"),
                            ("SIT", "SIT"),
                            ("ETC", "* PMS/EAP/완화연구/혈액,조직연구/기타")
                      ],
                      'PI': PI_list,
                      'monthly_enroll_by_PI_list': monthly_enroll_by_PI_list,
                      'from_date': from_date,
                      'to_date': to_date,
                      'date_year': today.strftime('%Y'),
                  })

@login_required
def ongoing_patients(request):
    pi = request.GET.get('pi') or ''
    type_label = request.GET.get('type') or ''
    crc = request.GET.get('crc') or ''

    # investigators
    investigators = InvestigatorContact.objects.filter(onco_A=True).values_list('name', flat=True)

    # ETC= 2개 이상 type
    type_counts = Research.objects.filter(is_deleted=0)\
        .values('id').annotate(type_count=Count('type', distinct=True))\
        .filter(type_count__gte=2).values_list('id', flat=True)
    etc_ids = set(type_counts)
    iit_ids = set(Research.objects.filter(is_deleted=0, type__value='IIT').values_list('id', flat=True))
    sit_ids = set(Research.objects.filter(is_deleted=0, type__value='SIT').values_list('id', flat=True))

    # ongoing ids (너 로직 그대로)
    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    ongo_condition1 = Feedback.objects.filter(cycle='1', day='1').values('assignment_id')
    EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values('assignment_id')
    ongoing_ids = Feedback.objects.filter(
        (Q(cycle='1', day='1', dosing_date__gte=from_date) & Q(cycle='EOT', dosing_date__lt=to_date)) |
        (Q(cycle='1', day='1', dosing_date__year=today.year) & Q(cycle='1', day='1', dosing_date__month=today.month)) |
        (Q(cycle='1', day='1', dosing_date__lt=from_date) & ~Q(assignment_id__in=EOT_assign))
    ).values_list('assignment_id', flat=True).distinct()

    qs = Assignment.objects.filter(
        is_deleted=False,
        id__in=ongoing_ids,
        research__PI__in=investigators,
        research__PI=pi,
    ).select_related('research', 'curr_crc')

    if crc != '__ALL__':
        qs = qs.filter(curr_crc__name=crc)

    if type_label == 'IIT':
        qs = qs.filter(research_id__in=iit_ids)
    elif type_label == 'SIT':
        qs = qs.filter(research_id__in=sit_ids)
    elif type_label == 'ETC':
        qs = qs.filter(research_id__in=etc_ids)
    else:
        return JsonResponse({'patients': []})

    data = list(qs.values(
        'id',
        'research__research_name',
        'register_number',
        'no',
        'name',
        'sex',
        'age',
        'PI',
        'curr_crc__name',
    ).order_by('research__research_name'))
    return JsonResponse({'patients': data})



def gsi_monthly_enroll_context(selected_year=None):
    """
    GSI팀 월별 등록 현황 (템플릿 + 엑셀 공용 데이터 생성)
    """
    if not selected_year:
        selected_year = date.today().year

    LINE_ORDER = ["periop", "line1", "line2", "line3", "solid", "N/A"]
    LINE_COLOR = {
        "periop": "#fff7e6",
        "adjuvant": "#fff7e6",
        "neoadjuvant": "#fff7e6",
        "line1": "#e6f7ff",
        "line2": "#e6ffe6",
        "line3": "#f9e6ff",
        "line4_or_more": "#fff0f6",
        "solid": "#e8e8e8",
        "na": "#f4f4f4",
        "etc": "#f4f4f4",
    }
    months = list(range(1, 13))

    # 포함 대상: Stomach, Sarcoma(전체), Urological(GSI팀만)
    stomach_ids = Image.objects.filter(cancer="Stomach").values_list("research_id", flat=True)
    sarcoma_ids = Image.objects.filter(cancer="Sarcoma").values_list("research_id", flat=True)
    urological_ids = Image.objects.filter(cancer="Urological", research__team="GSI").values_list("research_id", flat=True)

    research_ids = list(set(stomach_ids) | set(sarcoma_ids) | set(urological_ids))

    # 연구 기본 정보 (Line, 연구명)
    research_qs = Research.objects.filter(id__in=research_ids).prefetch_related("line")
    research_map = {}
    for r in research_qs:
        line_obj = r.line.first()
        if line_obj:
            line_value = getattr(line_obj, "value", None) or getattr(line_obj, "name", "N/A")
        else:
            line_value = "N/A"
        line_value = line_value.strip().lower()
        research_map[r.id] = {
            "name": r.research_name,
            "line": line_value,
            "line_display": dict(Line.CHOICES).get(line_value, line_value),
        }

    # Feedback 데이터
    feedback_qs = Feedback.objects.filter(assignment__research_id__in=research_ids, assignment__is_deleted=0)

    screening_qs = feedback_qs.filter(ICF_date__isnull=False, ICF_date__year=selected_year)
    enroll_qs = feedback_qs.filter(cycle="1", day="1", dosing_date__isnull=False, dosing_date__year=selected_year)

    screening_rows = (
        screening_qs.annotate(m=ExtractMonth("ICF_date"), rid=F("assignment__research_id"))
        .values("rid", "m")
        .annotate(c=Count("id"))
    )

    enroll_rows = (
        enroll_qs.annotate(m=ExtractMonth("dosing_date"), rid=F("assignment__research_id"))
        .values("rid", "m")
        .annotate(c=Count("id"))
    )

    # 연구별 월별 데이터 초기화
    per_research = {
        rid: {
            "name": research_map[rid]["name"],
            "line": research_map[rid]["line"],
            "line_display": research_map[rid]["line_display"],
            "months": {m: {"screening": 0, "enroll": 0} for m in months},
        }
        for rid in research_ids
    }

    for row in screening_rows:
        per_research[row["rid"]]["months"][row["m"]]["screening"] = row["c"]
    for row in enroll_rows:
        per_research[row["rid"]]["months"][row["m"]]["enroll"] = row["c"]

    # 정렬
    def line_key(line):
        return LINE_ORDER.index(line) if line in LINE_ORDER else 999

    sorted_items = sorted(
        per_research.items(),
        key=lambda it: (line_key(it[1]["line"]), it[1]["name"] or "")
    )

    # 라인별 그룹화
    line_groups = OrderedDict()
    for rid, data in sorted_items:
        line_groups.setdefault(data["line"], []).append((rid, data))

    # 테이블 렌더링용 구조 (템플릿과 동일)
    rows = []
    overall_month_sum = {m: {"screening": 0, "enroll": 0} for m in months}
    overall_total_screen = 0
    overall_total_enroll = 0

    for line, items in line_groups.items():
        line_month_sum = {m: {"screening": 0, "enroll": 0} for m in months}
        line_total_screen = 0
        line_total_enroll = 0

        for idx, (rid, data) in enumerate(items):
            row_total_screen = sum(data["months"][m]["screening"] for m in months)
            row_total_enroll = sum(data["months"][m]["enroll"] for m in months)

            for m in months:
                line_month_sum[m]["screening"] += data["months"][m]["screening"]
                line_month_sum[m]["enroll"] += data["months"][m]["enroll"]

            line_total_screen += row_total_screen
            line_total_enroll += row_total_enroll

            rows.append({
                "type": "data",
                "show_line": (idx == 0),
                "line": line,
                "line_display": data["line_display"],
                "line_rowspan": len(items),
                "line_color": LINE_COLOR.get(line, "#f4f4f4"),
                "rid": rid,
                "name": data["name"],
                "months": data["months"],
                "row_total_screen": row_total_screen,
                "row_total_enroll": row_total_enroll,
            })

        # 각 line 합계행
        rows.append({
            "type": "line_sum",
            "line": data["line_display"],
            "line_color": "#d9d9d9",
            "months": line_month_sum,
            "line_total_screen": line_total_screen,
            "line_total_enroll": line_total_enroll,
        })

        # 전체 합계 누적
        for m in months:
            overall_month_sum[m]["screening"] += line_month_sum[m]["screening"]
            overall_month_sum[m]["enroll"] += line_month_sum[m]["enroll"]

        overall_total_screen += line_total_screen
        overall_total_enroll += line_total_enroll

    return {
        "rows": rows,
        "months": months,
        "overall_month_sum": overall_month_sum,
        "overall_total_screen": overall_total_screen,
        "overall_total_enroll": overall_total_enroll,
        "selected_year": selected_year,
    }


@login_required()
def ETC_statistics(request):
    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    onco_A = Contact.objects.filter(onco_A=1).values('user_id')
    investigators = InvestigatorContact.objects.filter(onco_A=True).values_list('name', flat=True)

    # Status 불일치 명단
    ongo_condition1 = Feedback.objects.filter(cycle='1', day='1').values('assignment_id')
    EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values('assignment_id')
    status_discord_list = Feedback.objects.filter(~Q(assignment_id__in=EOT_assign) & Q(assignment__is_deleted=0, cycle='1', day='1') &
                                                  ~Q(assignment__status='ongoing') & ~Q(assignment__status='FU')) \
                                          .values('assignment_id').distinct() \
                                          .values('assignment__research__research_name', 'assignment__id', 'assignment__name', 'assignment__curr_crc__name')
    status_discord_2_list = Feedback.objects.filter(Q(assignment_id__in=EOT_assign, assignment__is_deleted=0) &
                                                    ~Q(assignment__status='off') & ~Q(assignment__status='FU')) \
                                            .values('assignment_id').distinct() \
                                            .values('assignment__research__research_name', 'assignment__id', 'assignment__name','assignment__curr_crc__name')

    # Total Visit Counts - 정규 분포
    total_visit_count_annotations = {}
    teams = ('TOTAL', 'GSI', 'CLUE')
    for team in teams:
        if team == 'GSI' or team == 'CLUE':
            total_visit_count_annotations[team] = Coalesce(Count('id', filter=(Q(uploader__contact__team__name=team, dosing_date__gte=from_date, dosing_date__lte=to_date)) & ~Q(cycle='')), 0)
        else:
            total_visit_count_annotations[team] = Coalesce(Count('id', filter=(Q(dosing_date__gte=from_date, dosing_date__lte=to_date)) & ~Q(cycle='')), 0)

        if Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A).count() == 0:
            total_visit = [{'TOTAL': 0, 'GSI': 0, 'CLUE': 0}]
        else:
            total_visit = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
                .values('uploader_id').distinct() \
                .annotate(**total_visit_count_annotations) \
                .values('uploader', *total_visit_count_annotations.keys())

    total_visit_dict = collections.defaultdict(list)
    for team_, count in itertools.product(teams, total_visit):
        total_visit_dict[team_].append(int(count[team_]))

    # Total Visit Counts - 정규 분포 - X축, series(data) 생성
    total_visit_count_normalization = collections.defaultdict(list)
    for team_ in teams:
        total_visit_max = max(total_visit_dict.get(team_))  # [11, 34, 55, 88] -> 88
        total_max_10 = total_visit_max if total_visit_max % 10 == 0 else total_visit_max + 10 - (total_visit_max % 10)  # 90
        total_visit_int_list = (list(int(n) for n in range(0, total_max_10 + 10, 10)))  # [0, 10, 20, 30, 40, 50, 60, 70, 80, 90]
        total_visit_count_normalization[f'{team_}_visit_xaxis_list'] += list(str(n) for n in range(0, total_max_10 + 10, 10))  # ['0', '10', '20', '30', '40', '50', '60', '70', '80', '90']
        total_visit_count_normalization[f'{team_}_visit_series_list'].append(list(sum(i < e <= i + 10 for e in total_visit_dict.get(team_)) for i in total_visit_int_list))  # 각 범위별 포함되는 요소 개수

    # 연구별 FU/EOS 수
    withdrawal_list = Research.objects.filter(is_deleted=0, onco_A=True, PI__in=investigators)\
                                 .values('id').distinct()\
                                 .annotate(death=Count('assignment', distinct=True, filter=Q(assignment__feedback__eos__value='death')),
                                           withdrawal=Count('assignment', distinct=True, filter=Q(assignment__feedback__eos__value='withdrawal')),
                                           etc=Count('assignment', distinct=True, filter=Q(assignment__feedback__eos__value='etc')),
                                           image_fu=Count('assignment', distinct=True, filter=Q(assignment__feedback__fu__value='image')),
                                           survival_fu=Count('assignment', distinct=True, filter=Q(assignment__feedback__fu__value='survival')),
                                           sum_EOS=F('death')+F('withdrawal')+F('etc'))\
                                 .values('id', 'research_name', 'is_recruiting', 'PI', 'death', 'withdrawal', 'etc', 'image_fu', 'survival_fu')\
                                 .order_by('-sum_EOS')

    withdrawal = {}
    for item in withdrawal_list:
        withdrawal.setdefault(item['PI'], []).append(item)

    selected_year = request.GET.get("year")
    if selected_year:
        selected_year = int(selected_year)
        tab = 'status01_04'
        context = gsi_monthly_enroll_context(selected_year)
    else:
        selected_year = date.today().year
        tab = ''
        context = gsi_monthly_enroll_context(selected_year)

    context.update({
        'today': today,
        'count': count,
        'from_date': from_date,
        'to_date': to_date,
        'date_year': today.strftime('%Y'),
        'total_visit_count_normalization': total_visit_count_normalization,
        'status_discord_list': status_discord_list,
        'status_discord_2_list': status_discord_2_list,
        'withdrawal': withdrawal,
        'teams': Team.objects.all(),
        'tab': tab,
    })

    return render(request, 'pages/research/statistics/ETC_statistics.html', context)

def generate_gsi_excel(context):
    """
    GSI 월별 등록 현황 엑셀 생성 (템플릿 구조 그대로 병합/색상 포함)
    """
    rows = context["rows"]
    months = context["months"]
    overall_month_sum = context["overall_month_sum"]
    overall_total_screen = context["overall_total_screen"]
    overall_total_enroll = context["overall_total_enroll"]
    selected_year = context["selected_year"]

    wb = Workbook()
    ws = wb.active
    ws.title = "GSI 월별 등록 현황"

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # 헤더 (2줄 구조)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.cell(1, 1, "Line").font = header_font
    ws.cell(1, 2, "Study Name").font = header_font

    col = 3
    for m in months:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(1, col, f"{m}월").font = header_font
        ws.cell(2, col, "Screening").font = header_font
        ws.cell(2, col + 1, "Enroll").font = header_font
        col += 2
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    ws.cell(1, col, "Total").font = header_font
    ws.cell(2, col, "Screening").font = header_font
    ws.cell(2, col + 1, "Enroll").font = header_font

    current_row = 3

    def rgb(color):
        return f"FF{color.replace('#', '')}"

    # 본문 (rows 반복)
    for r in rows:
        if r["type"] == "data":
            # Line 셀 병합
            if r["show_line"]:
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row + r["line_rowspan"] - 1, end_column=1)
                ws.cell(current_row, 1, r["line"].capitalize())
                ws.cell(current_row, 1).alignment = center_align
                ws.cell(current_row, 1).fill = PatternFill(start_color=rgb(r["line_color"]), fill_type="solid")

            # 연구명
            ws.cell(current_row, 2, r["name"])
            ws.cell(current_row, 1).fill = PatternFill(start_color=rgb(r["line_color"]).replace("#", ""), fill_type="solid")

            col = 3
            for m in months:
                md = r["months"][m]
                ws.cell(current_row, col, md["screening"])
                ws.cell(current_row, col + 1, md["enroll"])
                ws.cell(current_row, col).fill = PatternFill(start_color=r["line_color"].replace("#", ""), fill_type="solid")
                ws.cell(current_row, col + 1).fill = PatternFill(start_color=r["line_color"].replace("#", ""), fill_type="solid")
                col += 2

            ws.cell(current_row, col, r["row_total_screen"])
            ws.cell(current_row, col + 1, r["row_total_enroll"])
            ws.cell(current_row, col).fill = PatternFill(start_color=r["line_color"].replace("#", ""), fill_type="solid")
            ws.cell(current_row, col + 1).fill = PatternFill(start_color=r["line_color"].replace("#", ""), fill_type="solid")
            current_row += 1

        elif r["type"] == "line_sum":
            # 라인 합계 행
            ws.cell(current_row, 1, f"[{r['line']}] 합계")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(current_row, 1).fill = PatternFill(start_color="D9D9D9", fill_type="solid")

            col = 3
            for m in months:
                md = r["months"][m]
                ws.cell(current_row, col, md["screening"])
                ws.cell(current_row, col + 1, md["enroll"])
                col += 2
            ws.cell(current_row, col, r["line_total_screen"])
            ws.cell(current_row, col + 1, r["line_total_enroll"])
            for c in range(1, col + 2):
                ws.cell(current_row, c).fill = PatternFill(start_color="D9D9D9", fill_type="solid")
            current_row += 1

    # 전체 합계
    ws.cell(current_row, 1, "전체 합계")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    col = 3
    for m in months:
        md = overall_month_sum[m]
        ws.cell(current_row, col, md["screening"])
        ws.cell(current_row, col + 1, md["enroll"])
        col += 2
    ws.cell(current_row, col, overall_total_screen)
    ws.cell(current_row, col + 1, overall_total_enroll)
    for c in range(1, col + 2):
        ws.cell(current_row, c).fill = PatternFill(start_color="FFE8A6", fill_type="solid")

    # 전체 스타일
    for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1):
        for cell in row:
            cell.alignment = center_align
            if cell.row in (1, 2):
                cell.font = header_font

    ws.freeze_panes = "C3"

    # 응답으로 반환
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
            "attachment; filename*=UTF-8''" + quote(f"GSI_월별등록현황_{selected_year}.xlsx")
    )
    wb.save(response)
    return response

@login_required
def gsi_monthly_enroll_excel(request):
    selected_year = int(request.GET.get("year", date.today().year))
    context = gsi_monthly_enroll_context(selected_year)

    return generate_gsi_excel(context)


@login_required()
def not_entered_statistic(request):
    onco_A = Contact.objects.filter(onco_A=1).values('user_id')
    today = datetime.today()

    # 최근 일주일간 데이터 입력자
    data_entry_in_the_last_week = Feedback.objects.filter(assignment__is_deleted=0,
                                                          create_date__gte=today - timedelta(days=7),
                                                          create_date__lte=today).values('uploader_id').distinct()
    # 현 종양내과 A팀 & 데이터 입력자 제외 & group name이 1.nurse 2.medeical records 중 하나인 직원
    no_data_entry_in_the_last_week = Contact.objects.filter(Q(onco_A=1) & ~Q(user_id__in=data_entry_in_the_last_week)) \
                                                    .filter(Q(user_id__groups__name='nurse') | Q(user_id__groups__name='medical records'))

    # CRC별 담당 환자 평균 방문 일수 - 장기간 데이터 미입력 명단
    type_counts = Research.objects.filter(is_deleted=0).values('id').annotate(
        type_count=Count('type', distinct=True)).filter(type_count__gte=2).values('id')

    number_of_visits = Assignment.objects.filter(Q(is_deleted=0, research__is_deleted=0) & (
                Q(research__type__value='IIT') | Q(research__type__value='SIT')) & ~Q(research__in=type_counts)) \
        .annotate(number_of_visits=Count('feedback', filter=Q(feedback__dosing_date__isnull=False))) \
        .filter(number_of_visits__gte=2) \
        .values('id')

    avg_number_of_visits = Feedback.objects.filter(assignment__is_deleted=0, assignment__curr_crc__isnull=False) \
        .filter(
        Q(dosing_date__isnull=False, uploader__isnull=False, uploader__in=onco_A, assignment_id__in=number_of_visits) &
        ~Q(assignment__research__is_recruiting='Completed') & ~Q(
            assignment__research__is_recruiting='Not yet recruiting')) \
        .values('assignment') \
        .values('dosing_date').distinct() \
        .values('assignment', 'dosing_date', 'assignment__research__research_name', 'assignment__research__crc__name')

    if avg_number_of_visits.count() == 0:
        long_term_non_entry = []
    else:
        df_avg_number_of_visits = pd.DataFrame.from_records(avg_number_of_visits)
        # 환자별 평균 방문 일수 구하기
        assignment_avg = df_avg_number_of_visits.groupby(
            ['assignment', 'assignment__research__research_name', 'assignment__research__crc__name'])[
            "dosing_date"].apply(lambda x: x.diff().mean()).reset_index()
        assignment_avg['dosing_date'] = assignment_avg['dosing_date'].dt.days
        # 연구별 평균 방문 일수 구하기
        research_avg = round(
            assignment_avg.groupby(['assignment__research__research_name', 'assignment__research__crc__name'])[
                'dosing_date'].mean(), 2).reset_index()
        # CRC별 담당 환자 방문 일수 구하기
        crc_avg = round(
            research_avg.groupby('assignment__research__crc__name')['dosing_date'].mean()).reset_index().rename(
            columns={'assignment__research__crc__name': 'uploader__first_name'})
        # crc_avg['dosing_date'] = crc_avg['dosing_date'].astype('int')
        crc_avg['dosing_date'] = crc_avg['dosing_date'].fillna(0)

        latest_input = Feedback.objects.filter(assignment__is_deleted=0, assignment__curr_crc__isnull=False) \
            .filter(Q(dosing_date__isnull=False, uploader__isnull=False, uploader__in=onco_A) &
                    ~Q(assignment__research__is_recruiting='Completed') & ~Q(
            assignment__research__is_recruiting='Not yet recruiting')) \
            .values('uploader') \
            .annotate(latest_cycle=Max('create_date')) \
            .values('uploader__first_name', 'latest_cycle') \
            .order_by('uploader__first_name')

        df_latest_input = pd.DataFrame.from_records(latest_input)
        df_latest_input['latest_cycle'] = pd.to_datetime(df_latest_input['latest_cycle']).dt.date

        non_entry = pd.merge(df_latest_input, crc_avg, on="uploader__first_name", how="left").fillna(0)
        non_entry['min_data_entry_date'] = non_entry.latest_cycle + pd.to_timedelta(non_entry.dosing_date, unit='d')
        non_entry = non_entry[non_entry['min_data_entry_date'] < date.today()]
        non_entry[['latest_cycle', 'min_data_entry_date']] = non_entry[['latest_cycle', 'min_data_entry_date']].apply(
            lambda x: pd.to_datetime(x).dt.strftime('%Y-%m-%d'))

        json_non_entry = non_entry.to_json(orient='records')
        long_term_non_entry = json.loads(json_non_entry)

    # EOT 미입력 추적
    # 1. 연구별 첫 투약일로부터 EOT까지 평균 기간 (dict)
    ongo_condition1 = Feedback.objects.filter(cycle='1', day='1', dosing_date__isnull=False).values('assignment_id')
    EOT_notnull_assign = Feedback.objects.filter(
        Q(cycle='EOT', dosing_date__isnull=False) & Q(assignment_id__in=ongo_condition1)).values('assignment_id')

    C1D1_EOT_by_assignment = Feedback.objects.filter(Q(assignment_id__in=EOT_notnull_assign) &
                                                     (Q(assignment__research__type__value='IIT') | Q(
                                                         assignment__research__type__value='SIT')) & ~Q(
        assignment__research__in=type_counts)) \
        .values('assignment') \
        .annotate(C1D1_EOT_diff=ExpressionWrapper(
        Max('dosing_date', filter=Q(cycle='EOT')) - Max('dosing_date', filter=Q(cycle='1', day='1')),
        output_field=DurationField())) \
        .values('assignment', 'assignment__research__research_name', 'C1D1_EOT_diff')

    if C1D1_EOT_by_assignment.count() == 0:
        predict_EOT_date = []
    else:
        C1D1_EOT_by_assignment = pd.DataFrame.from_records(C1D1_EOT_by_assignment)
        C1D1_EOT_by_assignment['C1D1_EOT_diff'] = C1D1_EOT_by_assignment['C1D1_EOT_diff'].dt.days
        C1D1_EOT_by_research = C1D1_EOT_by_assignment.groupby(['assignment__research__research_name'])[
            'C1D1_EOT_diff'].mean()

        # json_C1D1_EOT_by_research = C1D1_EOT_by_research.to_json(orient='records')
        # avg_C1D1_EOT_by_research = json.loads(json_C1D1_EOT_by_research)

        # 2-1) status='OFF/FU' 이지만 EOT 미입력 or EOT 추가 안되어 있는 C1D1 환자의 C1D1 date + 연구별 EOT 평균 일수 < today
        C1D1 = Feedback.objects.filter(cycle='1', day='1').values('assignment_id')
        EOT = Feedback.objects.filter(cycle='EOT', assignment_id__in=C1D1).values('assignment_id')
        none_EOT = Feedback.objects.filter(
            Q(assignment__is_deleted=0, assignment__research__is_deleted=0) & Q(cycle='1', day='1') & ~Q(
                assignment_id__in=EOT)) \
            .annotate(C1D1_date=Max('dosing_date', filter=Q(cycle='1', day='1'))) \
            .values('assignment__name', 'assignment__curr_crc__name', 'assignment__id',
                    'assignment__research__research_name', 'C1D1_date').distinct() \
            .order_by('assignment__curr_crc__name')

        none_EOT = pd.DataFrame.from_records(none_EOT)
        none_EOT['C1D1_date'] = pd.to_datetime(none_EOT['C1D1_date']).dt.date
        none_EOT['C1D1_EOT_by_research'] = none_EOT['assignment__research__research_name'].map(C1D1_EOT_by_research)
        none_EOT = none_EOT.dropna(subset=['C1D1_EOT_by_research'])
        none_EOT['predict_EOT_date'] = none_EOT['C1D1_date'] + pd.to_timedelta(
            none_EOT['C1D1_EOT_by_research'].astype(int), unit='d')
        none_EOT = none_EOT[none_EOT['predict_EOT_date'] < date.today()]
        none_EOT[['C1D1_date', 'predict_EOT_date']] = none_EOT[['C1D1_date', 'predict_EOT_date']].apply(
            lambda x: pd.to_datetime(x).dt.strftime('%Y-%m-%d'))
        none_EOT = none_EOT.to_json(orient='records')
        none_EOT = json.loads(none_EOT)

        predict_EOT_date = {}
        for item in none_EOT:
            predict_EOT_date.setdefault(item['assignment__curr_crc__name'], []).append(item)

    # 한 달간 등록되지 않은 연구
    investigators = InvestigatorContact.objects.filter(onco_A=True).values_list('name', flat=True)
    stop_enroll_list = Research.objects.filter(is_deleted=0, is_recruiting='Recruiting', onco_A=True, PI__in=investigators) \
            .values('id') \
            .annotate(assignments=Sum(Case(When((Q(assignment__is_deleted=0) & ~Q(assignment__status='pre-screening') & ~Q(assignment__status='pre-screening-fail')), then=1), default=0, output_field=IntegerField())),
                      latest_enroll=Max('assignment__feedback__dosing_date', filter=Q(assignment__feedback__cycle='1', assignment__feedback__day='1'))) \
            .filter(Q(assignments=0) | Q(assignments__gte=1, latest_enroll__lte=date.today() - relativedelta.relativedelta(months=1))) \
            .values('PI', 'id', 'research_name', 'assignments', 'latest_enroll', 'create_date') \
            .order_by('PI', '-latest_enroll')

    stop_enroll = {}
    for item in stop_enroll_list:
        stop_enroll.setdefault(item['PI'], []).append(item)

    return render(request, 'pages/research/statistics/not_entered_statistics.html', {
        'today': today,
        'long_term_non_entry': long_term_non_entry,
        'predict_EOT_date': predict_EOT_date,
        'stop_enroll': stop_enroll,
        'no_data_entry_in_the_last_week': no_data_entry_in_the_last_week})        


@login_required()
def monthly_enroll(request):
    today = datetime.today()
    year = request.GET.get('year')
    month = request.GET.get('month')
    team = request.GET.get('team')

    monthly_enroll = Feedback.objects.filter(Q(assignment__is_deleted=0, assignment__research__is_deleted=0,
                                               cycle='1', day='1', dosing_date__year=year, dosing_date__month=month, uploader__contact__team__name=team)) \
        .values('dosing_date', 'assignment__research__research_name', 'assignment__PI', 'assignment__curr_crc__name', 'assignment__name', 'assignment__id', 'uploader__contact__team__name') \
        .order_by('dosing_date')

    # monthly_t_enroll = Supporting.objects.filter(Q(is_deleted=0, lab_date__year=year, lab_date__month=month) & ~Q(technician='')) \
    #     .annotate(technician_onco_A=Subquery(Contact.objects.filter(name=OuterRef('technician')).values('onco_A')[:1])) \
    #     .filter(technician_onco_A=True) \
    #     .values('lab_date', 'supporting_type', 'assignment__research__research_name', 'assignment__PI', 'assignment__name', 'technician') \

    return render(request, 'pages/research/statistics/monthly_enroll.html', {'monthly_enroll_list': monthly_enroll, 'today': today})


@login_required()
def monthly_visit(request):
    today = datetime.today()
    year = request.GET.get('year')
    month = request.GET.get('month')
    team = request.GET.get('team')

    monthly_visit = Feedback.objects.filter(Q(assignment__is_deleted=0, assignment__research__is_deleted=0,
                                              dosing_date__year=year, dosing_date__month=month, uploader__contact__team__name=team) & ~Q(cycle='')) \
                                    .prefetch_related('eos', 'fu')\
                                    .order_by('dosing_date')

    return render(request, 'pages/research/statistics/monthly_visit.html', {'monthly_visit_list': monthly_visit, 'today': today})


@login_required()
def statistic_supporting(request):
    return render(request, 'pages/research/statistic_supporting.html', {})


@login_required()
def download_ongoing(request):
    today = datetime.today()
    from_date = datetime(today.year, today.month, 1)
    to_date = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    date_year = today.year
    date_month = today.month
    ongo_condition1 = Feedback.objects.filter(cycle='1', day='1', dosing_date__lt=from_date).values('assignment_id')
    EOT_assign = Feedback.objects.filter(Q(cycle='EOT') & Q(assignment_id__in=ongo_condition1)).values('assignment_id')
    ongo_condition2 = Feedback.objects.filter(cycle='1', day='1', dosing_date__gte=from_date).values('assignment_id')
    EOT_assign2 = Feedback.objects.filter(
        Q(cycle='EOT', dosing_date__lt=to_date) & Q(assignment_id__in=ongo_condition2)).values('assignment_id')

    total_ongoing_condition = Feedback.objects.filter(assignment__is_deleted=False).filter(
        (Q(cycle='1', day='1', dosing_date__gte=from_date) & Q(assignment_id__in=EOT_assign2)) |
        (Q(cycle='1', day='1', dosing_date__year=date_year) & Q(cycle='1', day='1', dosing_date__month=date_month)) |
        # (Q(cycle='EOT', dosing_date__year=date_year) & Q(cycle='EOT', dosing_date__month=date_month)) |
        (Q(cycle='1', day='1', dosing_date__lt=from_date) & ~Q(assignment_id__in=EOT_assign))
    ).values('assignment_id').distinct()

    #observe_and_PMS = Research.objects.filter(
    #    Q(research_explanation__icontains='(관찰연구)') | Q(research_explanation__icontains='(PMS)') |
    #    Q(research_explanation__icontains='(완화연구)') | Q(research_explanation__icontains='(EAP)')).values('id')

    total_ongoing = Assignment.objects.filter(Q(research__is_deleted=0) & Q(id__in=total_ongoing_condition) &
                                              ~Q(curr_crc__isnull=True) & Q(curr_crc__in=Contact.objects.filter(onco_A=1).values('id'))) \
        .values('PI', 'curr_crc__name', 'register_number', 'name', 'research_id__research_name')

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # TODO Create total ongoing list header
    title = workbook.add_format({'border': 1, 'bottom': 2, 'align': 'center', 'bold': 1,'bg_color': '#BDBDBD', 'font_size': 13})
    border_center = workbook.add_format({'align': 'center', 'border': 1, 'bold': 1})
    ongoing_start = workbook.add_format({'top': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})
    ongoing_end = workbook.add_format({'bottom': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})

    # Create Assignment header
    ongoing_header = [["주치의", 10],
                      ["담당 CRC", None],
                      ["Research name", 20],
                      ["Patient name", 10],
                      ["Register number", None]]

    title_style = title
    worksheet.merge_range('A1:E1', '진행 중인 전체 환자 명단', title_style)

    row = 1
    col = 0
    for a_header, width in ongoing_header:
        worksheet.write(row, col, a_header, border_center)
        if width is not None:
            worksheet.set_column(col, col, width)
        col += 1

        for i, feedback in enumerate(total_ongoing):
            if i == 0:
                style = ongoing_start
            elif i == len(total_ongoing) - 1:
                style = ongoing_end

    row += 1
    col = 0
    for ongoing in total_ongoing:
        worksheet.write(row, col, ongoing['PI'], style)
        col += 1
        worksheet.write(row, col, ongoing['curr_crc__name'], style)
        col += 1
        worksheet.write(row, col, ongoing['research_id__research_name'], style)
        col += 1
        worksheet.write(row, col, ongoing['name'], style)
        col += 1
        worksheet.write(row, col, ongoing['register_number'], style)
        col += 1

        row += 1
        col = 0

    # Close the workbook before sending the data.
    workbook.close()

    # Rewind the buffer.
    output.seek(0)

    filename = '진행 중인 전체 환자 명단.xlsx'
    try:
        filename.encode('ascii')
        file_expr = 'filename="{}"'.format(filename)
    except UnicodeEncodeError:
        file_expr = "filename*=utf-8''{}".format(quote(filename))

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; ' + file_expr

    return response


@login_required()
def download_performance(request):
    if 'half_of_year' in request.GET:
        performance_year = request.GET.get('performance_year')
        performance_half_of_year = request.GET.get('half_of_year')
        if performance_half_of_year == 'first_half':
            from_date = datetime(int(performance_year), 1, 1)
            to_date = datetime(int(performance_year), 6, 30)
        elif performance_half_of_year == 'second_half':
            from_date = datetime(int(performance_year), 7, 1)
            to_date = datetime(int(performance_year), 12, 31)
    else:
        today = datetime.today()
        if today.month < 7:
            from_date = datetime(today.year, 1, 1)
            to_date = datetime(today.year, 6, 30)
        else:
            from_date = datetime(today.year, 7, 1)
            to_date = datetime(today.year, 11, 30)

    workers = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=from_date, initiation_date__lte=to_date, set_up__isnull=False) \
        .values('id') \
        .annotate(worker=Count('set_up')) \
        .values('id', 'worker')

    type_counts = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=from_date, initiation_date__lte=to_date, set_up__isnull=False) \
        .values('id') \
        .annotate(type_count=Count('type', distinct=True)).filter(type_count__gte=2) \
        .values('id')

    onco_A = Contact.objects.filter(onco_A=1).values('user_id')
    ICF_count = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
            .values('assignment') \
            .annotate(ICF_count=Count('id', filter=Q(ICF_date__isnull=False)),
                      last_ICF_date=Max('ICF_date')) \
            .filter(ICF_count__gte=2) \
            .filter(last_ICF_date__range=[from_date, to_date]) \
            .values('assignment')            

    observe_and_PMS = Research.objects.filter(Q(type__value='EAP') | Q(type__value='PMS') | Q(type__value='Palliative') | Q(type__value='Blood') | Q(type__value='ETC')).values('id')
    performance = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
            .values('uploader_id__first_name').distinct().order_by('uploader_id__first_name') \
            .annotate(order=Case(When(uploader__groups__name='nurse', then=Value(1)),
                                 When(uploader__groups__name='SETUP', then=Value(2)),
                                 When(uploader__groups__name='medical records', then=Value(3)), output_field=IntegerField()),
                      ICF=Count('assignment_id', filter=Q(ICF_date__isnull=False, ICF_date__gte=from_date, ICF_date__lte=to_date) &
                                             ~Q(assignment__status='pre-screening') & ~Q(assignment__status='pre-screening-fail') &
                                             ~Q(assignment__in=ICF_count) &
                                             ~Q(assignment_id__research__id__in=observe_and_PMS), distinct=True) * Value(10),
                      pre_ICF=Count('id', filter=Q(ICF_date__isnull=False, ICF_date__gte=from_date, ICF_date__lte=to_date) &
                                                 ~Q(assignment__in=ICF_count) &
                                                 (Q(assignment__status='pre-screening') | Q(assignment__status='pre-screening-fail') |
                                                  Q(assignment_id__research__id__in=observe_and_PMS))) * Value(3),
                      re_ICF=Count('assignment', distinct=True, filter=Q(assignment__in=ICF_count)),
                      cycle_visit=Count('id', filter=Q(dosing_date__gte=from_date, dosing_date__lte=to_date) & ~Q(cycle=''))) \
            .values_list('uploader__groups__name', 'uploader_id__first_name', 'ICF', 'pre_ICF', 're_ICF', 'cycle_visit', 'uploader_id')\
            .order_by('order')
    performance_list = [list(i)[0:7] for i in performance]

    performance_setup = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__gte=from_date, initiation_date__lte=to_date, set_up__isnull=False) \
        .values('set_up') \
        .annotate(IIT_setup=Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=1).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(120), IntegerField()) +
                            Cast(Count('id', filter=(Q(type__value='IIT', id__in=workers.filter(worker=2).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(60), IntegerField()) +
                            Cast(Count('id', filter=( Q(type__value='IIT', id__in=workers.filter(worker=3).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(40), IntegerField()),
                  SIT_setup=Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=1).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(30), IntegerField()) +
                            Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=2).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(15), IntegerField()) +
                            Cast(Count('id', filter=(Q(type__value='SIT', id__in=workers.filter(worker=3).values('id')) & ~Q(id__in=type_counts)), distinct=True) * Value(10), IntegerField()),
                  ETC_setup=Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                    (Q(id__in=workers.filter(worker=1).values('id'))), distinct=True) * Value(10), IntegerField()) +
                            Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                    (Q(id__in=workers.filter(worker=2).values('id'))), distinct=True) * Value(5), IntegerField()) +
                            Cast(Count('id', filter=(Q(id__in=type_counts) | (~Q(type__value='IIT') & ~Q(type__value='SIT') & ~Q(id__in=type_counts))) &
                                                    (Q(id__in=workers.filter(worker=3).values('id'))), distinct=True) * Value(3.3), IntegerField())) \
        .values('set_up__user_id', 'IIT_setup', 'SIT_setup', 'ETC_setup')

    feedback_uploader = Feedback.objects.filter(assignment__is_deleted=0, uploader_id__in=onco_A) \
                                        .values('uploader_id').distinct().values_list('uploader_id', flat=True)
    subject_of_performance = Contact.objects.filter(Q(onco_A=1) & (Q(user_id__groups__name='nurse') | Q(user_id__groups__name='medical records') | Q(user_id__groups__name='SETUP'))) \
                                            .values_list('user__id', flat=True)

    s = set(feedback_uploader)
    none_feedback_uploader = [x for x in subject_of_performance if x not in s]
    for n_f_u in none_feedback_uploader:
        name = User.objects.filter(id=n_f_u).values_list('groups__name', 'first_name', 'id')
        performance_list.append([name[0][0], name[0][1], 0, 0, 0, 0, name[0][2]])
        performance_list.sort()

    setting_new = {dct['set_up__user_id']: itemgetter('IIT_setup', 'SIT_setup', 'ETC_setup')(dct) for dct in performance_setup}
    uploader_ids = map(lambda x: x[-1], performance_list)
    performance_setup_list = [[i, *setting_new.get(i, (0, 0, 0))] for i in uploader_ids]

    performance_technician = Supporting.objects.filter(Q(is_deleted=0) & ~Q(technician='')) \
            .values('technician') \
            .annotate(PK_Sampling=Count('id', filter=Q(lab_date__gte=from_date, lab_date__lte=to_date + timedelta(days=1)))) \
            .values('technician', 'PK_Sampling')

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # TODO Create total ongoing list header
    title_style = workbook.add_format({'border': 1, 'bottom': 2, 'align': 'center', 'bold': 1, 'bg_color': '#BDBDBD', 'font_size': 13})
    border_center = workbook.add_format({'align': 'center', 'border': 1, 'bold': 1})
    performance_start = workbook.add_format({'top': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})
    performance_end = workbook.add_format({'bottom': 1, 'left': 1, 'right': 1, 'align': 'center', 'valign': 'vcenter'})

    # Create Assignment header
    performance_header = [["직급", 15],
                          ["Name", 10],
                          ["ICF x10", 10],
                          ["ICF (Pre/관찰연구) x3", 20],
                          ["ICF (Re) x1", 15],
                          ["Cycle Visit x1", 15],
                          ["IIT Set Up x120", 15],
                          ["SIT Set Up x30", 15],
                          ["기타 Set Up x10", 15],
                          ["Total", 10]]

    worksheet.merge_range('A1:J1', str(from_date.strftime('%Y-%m-%d')) + '~' + str(to_date.strftime('%Y-%m-%d')), title_style)

    row = 1
    col = 0
    for a_header, width in performance_header:
        worksheet.write(row, col, a_header, border_center)
        if width is not None:
            worksheet.set_column(col, col, width)
        col += 1

        for i, feedback in enumerate(performance_list):
            if i == 0:
                style = performance_start
            elif i == len(performance_list) - 1:
                style = performance_end

    row += 1
    col = 0
    for p in performance_list:
        worksheet.write(row, col, p[0], style)
        col += 1
        worksheet.write(row, col, p[1], style)
        col += 1
        worksheet.write(row, col, p[2], style)
        col += 1
        worksheet.write(row, col, p[3], style)
        col += 1
        worksheet.write(row, col, p[4], style)
        col += 1
        worksheet.write(row, col, p[5], style)
        col += 1
        for p_s in performance_setup_list:
            if p_s[0] == p[6]:
                worksheet.write(row, col, p_s[1], style)
                col += 1
                worksheet.write(row, col, p_s[2], style)
                col += 1
                worksheet.write(row, col, p_s[3], style)
                col += 1
                worksheet.write(row, col, p[2] + p[3] + p[4] + p[5] + p_s[1] + p_s[2] + p_s[3], style)
        col += 1

        row += 1
        col = 0

        p_df = pd.DataFrame.from_records(performance_list)
        for p in p_df[0].unique():
            u = p_df.loc[p_df[0] == p].index.values + 1

            if len(u) < 2:
                pass
            else:
                merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter'})
                worksheet.merge_range(u[1], 0, u[-1] + 1, 0, p_df.loc[u[0], 0], merge_format)

    row += 1
    performance_technician_header = [["직급", 15], ["Name", 10], ["건수", 10]]

    for a_header, width in performance_technician_header:
        worksheet.write(row, col, a_header, border_center)
        if width is not None:
            worksheet.set_column(col, col, width)
        col += 1

        for i, feedback in enumerate(performance_technician):
            if i == 0:
                style = performance_start
            elif i == len(performance_technician) - 1:
                style = performance_end

    row += 1
    col = 0

    for i, p_t in enumerate(performance_technician):
        if i == 0:
            worksheet.merge_range(row, col, row + 1, col, 'technician', style)
            col += 1
            worksheet.write(row, col, p_t['technician'], style)
            col += 1
            worksheet.write(row, col, p_t['PK_Sampling'], style)
            col += 1
        elif i == 1:
            worksheet.write(row, col + 1, p_t['technician'], style)
            col += 1
            worksheet.write(row, col + 1, p_t['PK_Sampling'], style)
            col += 1

        row += 1
        col = 0

    # Close the workbook before sending the data. and Rewind the buffuer
    workbook.close()
    output.seek(0)

    filename = 'Performance.xlsx'
    try:
        filename.encode('ascii')
        file_expr = 'filename="{}"'.format(filename)
    except UnicodeEncodeError:
        file_expr = "filename*=utf-8''{}".format(quote(filename))

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; ' + file_expr

    performance_download = DownloadLog(downloader=request.user, content='Performance')
    performance_download.save()

    return response


@login_required()
@require_http_methods(['GET', 'POST'])
def add_waiting(request, cancer_id):
    cancer = get_object_or_404(Cancer, pk=cancer_id)
    parameter = request.GET.get('cancer', None)
    if request.method == 'GET':
        # Start creating new waitinglist
        return render(request, 'pages/waitinglist/add.html',
                      {
                       'parameter': parameter,
                       'cancer': cancer,
                       'waitinglist_field_choice': WaitingList.field_value_and_text()
                      })

    temp_waitinglist, errors = WaitingList.waitinglist_form_validation(request, cancer)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/waitinglist/add.html',
            {
                'parameter': parameter,
                'cancer': cancer,
                'waitinglist': temp_waitinglist,
                'editable': True,
                'waitinglist_field_choice': WaitingList.field_value_and_text(),
                'errors': error_msg
            }
        )

    # Create new waiting patient
    waitinglist = WaitingList(**dict(vars(temp_waitinglist)))
    waitinglist.save()
    # Redirect to the new assignment detail page.
    return HttpResponseRedirect('/research/waitinglist/' + str(waitinglist.id) + '/')

@login_required()
@require_http_methods(['GET', 'POST'])
def phase_add_waiting(request, phase_id):
    phase = get_object_or_404(Phase, pk=phase_id)
    parameter = request.GET.get('phase', None)
    if request.method == 'GET':
        # Start creating new waitinglist
        return render(request, 'pages/waitinglist/phase_add.html',
                      {
                       'parameter': parameter,
                       'phase': phase,
                       'waitinglist_field_choice': WaitingList.field_value_and_text()
                      })

    temp_waitinglist, errors = WaitingList.phase_waitinglist_form_validation(request, phase)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/waitinglist/phase_add.html',
            {
                'parameter': parameter,
                'phase': phase,
                'waitinglist': temp_waitinglist,
                'editable': True,
                'waitinglist_field_choice': WaitingList.field_value_and_text(),
                'errors': error_msg
            }
        )

    # Create new waiting patient
    waitinglist = WaitingList(**dict(vars(temp_waitinglist)))
    waitinglist.save()
    # Redirect to the new assignment detail page.
    return HttpResponseRedirect('/research/waitinglist/' + str(waitinglist.id) + '/')


@login_required()
def detail_waiting(request, waitinglist_id):
    waitinglist = get_object_or_404(WaitingList, pk=waitinglist_id)

    if waitinglist.is_deleted:
        return render(request, 'pages/waitinglist/deleted.html', {'waitinglist': waitinglist})

    return render(request, 'pages/waitinglist/detail.html',
                  {
                'waitinglist': waitinglist,
                'waitinglist_field_choice': WaitingList.field_value_and_text()
                  })


@login_required()
def edit_waiting(request, waitinglist_id):
    waitinglist = get_object_or_404(WaitingList, pk=waitinglist_id)
    edited_waitinglist, errors = WaitingList.waitinglist_form_validation(request, None)

    if errors:
        return JsonResponse({'code': 'error', 'error': errors})

    waitinglist.register_number = edited_waitinglist.register_number
    waitinglist.name = edited_waitinglist.name
    waitinglist.doctor = edited_waitinglist.doctor
    waitinglist.sex = edited_waitinglist.sex
    waitinglist.age = edited_waitinglist.age
    waitinglist.curr_status = edited_waitinglist.curr_status

    waitinglist.save()
    return JsonResponse({'code': 'success'})


@login_required()
def delete_waiting(request, waitinglist_id):
    waitinglist = get_object_or_404(WaitingList, pk=waitinglist_id)
    #cancer_value = waitinglist.cancer.value
    waitinglist.is_deleted = True
    waitinglist.delete()
    return HttpResponseRedirect(f'/research/search/')


@login_required()
@require_http_methods(['GET', 'POST'])
def add_research_waiting(request, research_id):
    research = get_object_or_404(Research, pk=research_id)
    parameter = request.GET.get('research', None)
    if request.method == 'GET':
        return render(request, 'pages/waitinglist/research_add.html',
                      {
                       'parameter': parameter,
                       'research': research,
                       'waitinglist_field_choice': WaitingList.field_value_and_text()
                      })

    temp_waitinglist, errors = research_WaitingList.research_waitinglist_form_validation(request, research)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/waitinglist/research_add.html',
            {
                'parameter': parameter,
                'research': research,
                'waitinglist': temp_waitinglist,
                'editable': True,
                'waitinglist_field_choice': research_WaitingList.field_value_and_text(),
                'errors': error_msg
            }
        )

    # Create new waiting patient
    research_waitinglist = research_WaitingList(**dict(vars(temp_waitinglist)))
    research_waitinglist.save()
    # Redirect to the new assignment detail page.
    research_id = research_waitinglist.research.id
    return HttpResponseRedirect(f'/research/{research_id}')


@login_required()
def edit_research_waiting(request, research_waitinglist_id):
    research_waitinglist = get_object_or_404(research_WaitingList, pk=research_waitinglist_id)
    edited_research_waitinglist, errors = research_WaitingList.research_waitinglist_form_validation(request, None)

    if errors:
        return JsonResponse({'code': 'error', 'error': errors})

    research_waitinglist.register_number = edited_research_waitinglist.register_number
    research_waitinglist.name = edited_research_waitinglist.name
    research_waitinglist.doctor = edited_research_waitinglist.doctor
    research_waitinglist.sex = edited_research_waitinglist.sex
    research_waitinglist.age = edited_research_waitinglist.age
    research_waitinglist.curr_status = edited_research_waitinglist.curr_status

    research_waitinglist.save()
    return JsonResponse({'code': 'success'})


@login_required()
def delete_research_waiting(request, research_waitinglist_id):
    research_waitinglist = get_object_or_404(research_WaitingList, pk=research_waitinglist_id)
    research_id = research_waitinglist.research.id
    research_waitinglist.is_deleted = True
    research_waitinglist.save()
    return HttpResponseRedirect(f'/research/{research_id}')


@login_required()
@require_http_methods(['GET', 'POST'])
def add_waiting_PI(request):
    if request.method == 'GET':
        # Start creating new waitinglist
        return render(request, 'pages/waitinglist/add_PI.html',
                      {
                       'waitinglist_field_choice': WaitingList.field_value_and_text()
                      })

    temp_waitinglist, errors = WaitingList.waitinglist_PI_form_validation(request)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/waitinglist/add_PI.html',
            {
                'waitinglist': temp_waitinglist,
                'editable': True,
                'waitinglist_field_choice': WaitingList.field_value_and_text(),
                'errors': error_msg
            }
        )

    # Create new waiting patient
    waitinglist = WaitingList(**dict(vars(temp_waitinglist)))
    waitinglist.save()
    # Redirect to the new assignment detail page.
    return HttpResponseRedirect('/research/waitinglist/' + str(waitinglist.id) + '/')


@login_required()
def pre_initiation(request):
    today = datetime.today()
    today_str = today.strftime('%Y-%m-%d')
    the_first_day_of_the_year = datetime(today.year, 1, 1)

    # EAP, PMS, 완화 연구 등
    type_counts = Pre_Initiation.objects.filter(is_deleted=0, initiation_date__year=today.year) \
                                        .values('id') \
                                        .annotate(type_count=Count('type', distinct=True)).filter(type_count__gte=2) \
                                        .values('id')

    dashboard = Pre_Initiation.objects.values('is_deleted')\
                        .annotate(scheduled_to_start_SIT=Count('id', distinct=True, filter=(Q(initiation_date__gt=today) & ~Q(Q(type__value='IIT') & ~Q(id__in=type_counts)))),
                                  scheduled_to_start_IIT=Count('id', distinct=True, filter=(Q(initiation_date__gt=today) & (Q(type__value='IIT') & ~Q(id__in=type_counts)))),
                                  withhold=Count('id', distinct=True, filter=Q(is_withhold='Y')),
                                  start_completion=Count('id', distinct=True, filter=Q(initiation_date__year=today.year, initiation_date__lte=today)))\
                        .filter(is_deleted=0)

    pre_initiation = Pre_Initiation.objects.filter(is_deleted=False).prefetch_related('set_up', 'cancer', 'phase', 'pre_initiation_sit_set', 'type').order_by('-initiation_date')
    implement_date = date(2022, 10, 20)

    return render(request, 'pages/research/pre_initiation/pre_initiation.html',
                        {'pre_initiation': pre_initiation, 'today': today_str, 'implement_date': implement_date, 'dashboard': dashboard,
                         'the_first_day_of_the_year': the_first_day_of_the_year,
                         'teams': Team.objects.all()})

@login_required()
def add_pre_initiation(request):
    if request.method == 'GET':
        new_pre_initiation = Pre_Initiation()
        return render(
            request, 'pages/research/pre_initiation/add.html', {
                'type': request.GET.get('type', ''),
                'pre_initiation': new_pre_initiation,
                'editable': True,
                'field_choice': Research.create_field_value_and_text(),
                'contact_value_and_text': Research.contact_value_and_text(),
                'teams': Team.objects.all()})

    temp_pre_initiation, errors = Pre_Initiation.pre_initiation_form_validation(request)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/research/pre_initiation/add.html',
            {
                'pre_initiation': temp_pre_initiation,
                'editable': True,
                'field_choice': Research.create_field_value_and_text(),
                'contact_value_and_text': Research.contact_value_and_text(),
                'errors': error_msg,
                'teams': Team.objects.all()
            }
        )

    field_dict = dict(vars(temp_pre_initiation))
    field_dict.pop('set_up')
    field_dict.pop('cancer')
    field_dict.pop('phase')
    #field_dict.pop('chemotherapy')
    #field_dict.pop('alternation')
    field_dict.pop('type')
    new_pre_initiation = Pre_Initiation(**field_dict)
    new_pre_initiation.save()
    new_pre_initiation.set_up.set(temp_pre_initiation.set_up)
    new_pre_initiation.cancer.set(temp_pre_initiation.cancer)
    new_pre_initiation.phase.set(temp_pre_initiation.phase)
    #new_pre_initiation.chemotherapy.set(temp_pre_initiation.chemotherapy)
    #new_pre_initiation.alternation.set(temp_pre_initiation.alternation)
    new_pre_initiation.type.set(temp_pre_initiation.type)

    return HttpResponseRedirect('/research/pre_initiation/')

@login_required()
def detail_pre_initiation(request, id):
    pre_initiation = Pre_Initiation.objects.filter(pk=id).prefetch_related('set_up', 'cancer', 'phase')
    pre_initiation = get_object_or_404(pre_initiation)

    study_categories = Study_Category.objects.all()
    study_memo = Study_Memo.objects.filter(pre_initiation_id=pre_initiation.id)

    iit_setup = pre_initiation.pre_initiation_iit_set.values('id', 'preperation__value', 'mfds__value', 'irb__value', 'crms__value', 
                                                             'multicenter__value', 'etc__value', 'from_date', 'to_date', 'memo')
    sit_setup = pre_initiation.pre_initiation_sit_set.values('id', 'feasibility', 'PSV', 'budgeting_from', 'budgeting_to', 'IRB_new_review', 
                                                             'IRB_qualified_permission', 'IRB_finalization', 'contract', 'SIV')
    
    onco_A_crc = Contact.objects.filter(onco_A=1).values_list('user_id', flat=True)
    onco_A_investigator = InvestigatorContact.objects.filter(onco_A=1).values_list('user_id', flat=True)
    onco = onco_A_crc.union(onco_A_investigator)

    return render(request, 'pages/research/pre_initiation/detail.html',
                  {'pre_initiation': pre_initiation, 'sit_setup': sit_setup, 'iit_setup': iit_setup,
                   'study_categories': study_categories,
                   'study_memo': serializers.serialize("json", study_memo),
                   'contact_value_and_text': Research.contact_value_and_text(),
                   'field_choice': Pre_Initiation_IIT.create_field_value_and_text(),
                   'onco': onco,
                   'editable': True})

@login_required()
def edit_pre_initiation(request, id):
    pre_initiation = Pre_Initiation.objects.filter(is_deleted=False, pk=id).prefetch_related('set_up', 'cancer', 'phase', 'type')
    pre_initiation = get_object_or_404(pre_initiation)
    crc = Contact.objects.filter(Q(onco_A=1) & ~Q(team__name='etc')).order_by('name')

    if pre_initiation.is_deleted:
        return HttpResponse(status=403, content='이미 삭제된 연구입니다.')

    # GET req
    if request.method == 'GET':
        return render(
            request, 'pages/research/pre_initiation/edit.html',
            {
             'crc': crc,
             'pre_initiation': pre_initiation,
             'editable': True,
             'field_choice': Research.create_field_value_and_text(),
             'contact_value_and_text': Research.contact_value_and_text(),
             'teams': Team.objects.all()
             }
        )

    temp_pre_initiation, errors = Pre_Initiation.pre_initiation_form_validation(request)
    crc = Contact.objects.filter(Q(onco_A=1) & ~Q(team__name='etc')).order_by('name')

    if errors:
        # Temporary set attributes. cf. It is not a model instance!
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        temp_pre_initiation.id = pre_initiation.id
        return render(
            request, 'pages/research/pre_initiation/edit.html',
            {
                'crc': crc,
                'pre_initiation': temp_pre_initiation,
                'editable': True,
                'field_choice': Research.create_field_value_and_text(),
                'errors': error_msg,
                'contact_value_and_text': Research.contact_value_and_text(),
                'teams': Team.objects.all()
            }
        )

    pre_initiation.team = temp_pre_initiation.team
    pre_initiation.study_code = temp_pre_initiation.study_code
    pre_initiation.pre_research_name = temp_pre_initiation.pre_research_name
    pre_initiation.study_explanation = temp_pre_initiation.study_explanation
    pre_initiation.PI = temp_pre_initiation.PI
    pre_initiation.crc = temp_pre_initiation.crc
    pre_initiation.is_withhold = temp_pre_initiation.is_withhold
    pre_initiation.is_commence = temp_pre_initiation.is_commence
    pre_initiation.sponsor = temp_pre_initiation.sponsor
    pre_initiation.CRO = temp_pre_initiation.CRO
    pre_initiation.initiation_date = temp_pre_initiation.initiation_date
    pre_initiation.CTC_contract = temp_pre_initiation.CTC_contract
    pre_initiation.CTC_non_contract_reason = temp_pre_initiation.CTC_non_contract_reason
    pre_initiation.memo = temp_pre_initiation.memo
    pre_initiation.tx = temp_pre_initiation.tx

    pre_initiation.set_up.set(temp_pre_initiation.set_up)
    pre_initiation.cancer.set(temp_pre_initiation.cancer)
    pre_initiation.phase.set(temp_pre_initiation.phase)
    #pre_initiation.chemotherapy.set(temp_pre_initiation.chemotherapy)
    #pre_initiation.alternation.set(temp_pre_initiation.alternation)
    pre_initiation.type.set(temp_pre_initiation.type)
    pre_initiation.save()

    return HttpResponseRedirect(f'/research/pre_initiation/{pre_initiation.id}/')

@login_required()
def delete_pre_initiation(request, id):
    pre_initiation = Pre_Initiation.objects.get(pk=id)
    pre_initiation.is_deleted = True
    pre_initiation.save()
    return HttpResponseRedirect(f'/research/pre_initiation/')


@login_required()
@require_http_methods(['GET', 'POST'])
def add_SIT_setup(request, pre_initiation_id):
    pre_initiation = get_object_or_404(Pre_Initiation, pk=pre_initiation_id)
    _setup, errors = Pre_Initiation_SIT.SIT_setup_form_validation(request, pre_initiation)

    if errors:
        return JsonResponse({'code': 'error', 'error': errors})

    field_dict = dict(vars(_setup))
    new_setup = Pre_Initiation_SIT(**field_dict)
    new_setup.save()

    return JsonResponse({'code': 'success'})

@login_required()
def edit_SIT_setup(request, setup_id):
    setup = get_object_or_404(Pre_Initiation_SIT, pk=setup_id)
    edited_setup, errors = Pre_Initiation_SIT.SIT_setup_form_validation(request, None)

    if errors:
        return JsonResponse({'code': 'error', 'error': errors})

    setup.feasibility = edited_setup.feasibility
    setup.PSV = edited_setup.PSV
    setup.budgeting_from = edited_setup.budgeting_from
    setup.budgeting_to = edited_setup.budgeting_to
    setup.IRB_new_review = edited_setup.IRB_new_review
    setup.IRB_qualified_permission = edited_setup.IRB_qualified_permission
    setup.IRB_finalization = edited_setup.IRB_finalization
    setup.contract = edited_setup.contract
    setup.SIV = edited_setup.SIV
    setup.uploader = edited_setup.uploader
    setup.save()

    return JsonResponse({'code': 'success'})

@login_required()
def delete_SIT_setup(request, setup_id):
    setup = Pre_Initiation_SIT.objects.get(pk=setup_id)
    initiation_id = setup.pre_initiation.id
    setup.delete()
    return HttpResponseRedirect(f'/research/pre_initiation/{initiation_id}/')


@login_required()
@require_http_methods(['GET', 'POST'])
def add_IIT_setup(request, pre_initiation_id):
    pre_initiation = get_object_or_404(Pre_Initiation, pk=pre_initiation_id)
    _setup, errors = Pre_Initiation_IIT.IIT_setup_form_validation(request, pre_initiation)

    if errors:
        return JsonResponse({'code': 'error', 'error': errors})

    field_dict = dict(vars(_setup))
    new_setup = Pre_Initiation_IIT(**field_dict)
    new_setup.save()

    return JsonResponse({'code': 'success'})

@login_required()
@require_http_methods(['POST'])
def add_IIT_memo(request):
    pre_initiation_id = request.POST.get('pre_initiation_id')
    subcategory = request.POST.get('subcategory')
    from_date = request.POST.get('from_date')
    to_date = request.POST.get('to_date')
    memo = request.POST.get('memo')
    memo_id = request.POST.get('memo_id')

    try:
        from_date = datetime.strptime(from_date, '%m/%d/%Y')
    except:
        from_date = None
    try:
        to_date = datetime.strptime(to_date, '%m/%d/%Y')
    except:
        to_date = None

    if memo_id is None or memo_id == "":
        Study_Memo.objects.create(pre_initiation_id=pre_initiation_id, sub_category_id=subcategory,
                                  start_date=from_date, end_date=to_date, memo=memo)
    else:
        study_memo = Study_Memo.objects.get(id=memo_id)
        study_memo.memo = memo
        study_memo.start_date = from_date
        study_memo.end_date = to_date
        study_memo.save()
    return HttpResponseRedirect(f'/research/pre_initiation/{pre_initiation_id}')

@login_required()
def edit_IIT_setup(request, setup_id):
    setup = get_object_or_404(Pre_Initiation_IIT, pk=setup_id)
    edited_setup, errors = Pre_Initiation_IIT.IIT_setup_form_validation(request, None)

    if errors:
        return JsonResponse({'code': 'error', 'error': errors})

    setup.preperation = edited_setup.preperation
    setup.mfds = edited_setup.mfds
    setup.irb = edited_setup.irb
    setup.crms = edited_setup.crms
    setup.multicenter = edited_setup.multicenter
    setup.etc = edited_setup.etc
    setup.from_date = edited_setup.from_date
    setup.to_date = edited_setup.to_date
    setup.memo = edited_setup.memo
    setup.uploader = edited_setup.uploader
    setup.save()

    return JsonResponse({'code': 'success'})

@login_required()
def delete_IIT_setup(request, setup_id):
    setup = Pre_Initiation_IIT.objects.get(pk=setup_id)
    initiation_id = setup.pre_initiation.id
    setup.delete()
    return HttpResponseRedirect(f'/research/pre_initiation/{initiation_id}/')

@login_required()
@require_http_methods(['GET', 'POST'])
def update_research(request, id):
    pre_initiation = get_object_or_404(Pre_Initiation, pk=id)
    pre_initiation.is_deleted = False
    pre_initiation.save()

    if request.method == 'GET':
        new_research = Research()
        backup = Contact.objects.filter(onco_A=1).order_by('name')
        return render(
            request, 'pages/research/add.html', {
                'research': new_research,
                'backup': backup,
                'editable': True,
                'field_choice': Research.create_field_value_and_text(),
                'pre_initiation': pre_initiation,
                'contact_value_and_text': Research.contact_value_and_text(),
                'teams': Team.objects.all()
            }
        )

    temp_research, errors = Research.research_form_validation(request)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/research/add.html',
            {'research': temp_research,
             'editable': True,
             'field_choice': Research.create_field_value_and_text(),
             'errors': error_msg,
             'pre_initiation': pre_initiation,
             'contact_value_and_text': Research.contact_value_and_text(),
             'teams': Team.objects.all()
             }
        )

    field_dict = dict(vars(temp_research))
    field_dict.pop('cancer')
    field_dict.pop('phase')
    #field_dict.pop('alternation')
    #field_dict.pop('chemotherapy')
    new_research = Research(**field_dict)
    new_research.save()

    new_research.cancer.set(temp_research.cancer)
    new_research.phase.set(temp_research.phase)
    #new_research.alternation.set(temp_research.alternation)
    #new_research.chemotherapy.set(temp_research.chemotherapy)

    field_summary = compare_research_fields({}, new_research.json())
    file_summary = collections.defaultdict(list)

    # 연구 프로토콜 파일 (국문)
    if request.FILES.getlist('file[]'):
        file_summary['prev'] = None

    for _file in request.FILES.getlist('file[]'):
        new_file = UploadFile(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 연구 프로토콜 파일 (영문)
    if request.FILES.getlist('eng_file[]'):
        file_summary['prev'] = None

    for _file in request.FILES.getlist('eng_file[]'):
        new_file = UploadEngFile(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 주 선정기준 이미지
    if request.FILES.get('inclusion'):
        file_summary['prev'] = None

    for _file in request.FILES.get('inclusion'):
        new_file = UploadInclusion(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # 주 제외기준 이미지
    if request.FILES.get('exclusion'):
        file_summary['prev'] = None

    for _file in request.FILES.get('exclusion'):
        new_file = UploadExclusion(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    # Reference
    if request.FILES.get('reference'):
        file_summary['prev'] = None

    for _file in request.FILES.get('reference'):
        new_file = UploadReference(filename=_file.name, file=_file, research=new_research)
        new_file.save()
        file_summary['curr'].append(new_file.json())

    history = History(user=request.user,
                      research=new_research,
                      summary={
                          'field_summary': field_summary,
                          'file_summary': file_summary
                      },
                      history_type=History.CREATE_RESEARCH,
                      content=new_research.json())
    history.save()

    return HttpResponseRedirect('/research/' + str(research.id) + '/')


# 1. end_study ('Research' Class Model에서 끌고 온 데이터) 2. end_research (별도 추가)
@login_required()
def end_study(request):
    end_study = Research.objects.filter(is_deleted=0).filter(Q(status='종료보고완료')|Q(status='결과보고완료')|Q(status='장기보관완료')) \
        .annotate(model_type=Case(
                When(id__in=Research.objects.filter(is_deleted=0).values_list('id', flat=True), then=Value('end_study')), output_field=CharField()))\
        .values('id', 'research_name', 'study_code', 'PI', 'status', 'end_brief', 'result_brief', 'storage_date', 'model_type')\
        .order_by('PI', 'research_name')

    end_research = End_research.objects.filter(is_deleted=0) \
        .annotate(model_type=Case(
            When(id__in=End_research.objects.filter(is_deleted=0).values_list('id', flat=True), then=Value('end_research')), output_field=CharField())) \
        .values('id', 'research_name', 'study_code', 'PI', 'status', 'binder_location', 'study_coordinator', 'storage_date', 'end_brief', 'result_brief', 'sponsor', 'CRA_name', 'CRA_phoneNumber', 'model_type')\
        .order_by('PI', 'research_name')

    combined_queryset = list(chain(end_study, end_research))
    end = {}
    for item in combined_queryset:
        end.setdefault(item['PI'], []).append(item)

    return render(request, 'pages/research/end_study/end_study.html', {'end_study': end_study,
                                                                       'end_research': end_research,
                                                                       'end': end,
                                                                       'upload_images': UploadImage.objects.filter(is_deleted=False),
                                                                       'end_upload_images': End_UploadImage.objects.filter(is_deleted=False),
                                                                       'research_archives': Research_Archive.objects.filter(is_deleted=False),
                                                                       'end_research_archives': End_Research_Archive.objects.filter(is_deleted=False)})


@login_required()
@require_http_methods(['GET', 'POST'])
def add_end_research(request):
    if request.method == 'GET':
        return render(request, 'pages/research/end_study/add.html', {'status': request.GET.get('status', '')})

    temp_end_research, errors = End_research.end_research_form_validation(request)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/research/end_study/add.html',
            {
                'end_research': temp_end_research,
                'editable': True,
                'errors': error_msg
            }
        )

    end_research = End_research(**dict(vars(temp_end_research)))
    end_research.save()

    # 연구 바인더 위치 (사진)
    for _image in request.FILES.getlist('image[]'):
        new_image = End_UploadImage(imagename=_image.name, image=_image, end_research=end_research)
        new_image.save()

    # 장기 보관 문서 (파일)
    for _file in request.FILES.getlist('end_research_archive[]'):
        new_file = End_Research_Archive(filename=_file.name, file=_file, end_research=end_research)
        new_file.save()

    return HttpResponseRedirect('/research/end_study/')

@login_required()
def edit_end_research(request, id):
    end_research = get_object_or_404(End_research, pk=id)

    # GET req
    if request.method == 'GET':
        return render(
            request, 'pages/research/end_study/edit.html',
            {
             'end_research': end_research,
             'editable': True,
             'end_upload_images': End_UploadImage.objects.filter(is_deleted=False, end_research=end_research),
             'end_research_archives': End_Research_Archive.objects.filter(is_deleted=False, end_research=end_research)
             }
        )

    temp_end_research, errors = End_research.end_research_form_validation(request)

    if errors:
        error_msg = {k: '<br>'.join(v) for k, v in errors.items()}
        return render(
            request, 'pages/research/end_study/edit.html',
            {
                'end_research': temp_end_research,
                'editable': True,
                'errors': error_msg,
                'end_upload_images': End_UploadImage.objects.filter(is_deleted=False, end_research=end_research),
                'end_research_archives': End_Research_Archive.objects.filter(is_deleted=False, end_research=end_research)
            }
        )

    end_research.research_name = temp_end_research.research_name
    end_research.study_code = temp_end_research.study_code
    end_research.PI = temp_end_research.PI
    end_research.status = temp_end_research.status
    end_research.binder_location = temp_end_research.binder_location
    end_research.study_coordinator = temp_end_research.study_coordinator
    end_research.storage_date = temp_end_research.storage_date
    end_research.end_brief = temp_end_research.end_brief
    end_research.result_brief = temp_end_research.result_brief
    end_research.sponsor = temp_end_research.sponsor
    end_research.CRA_name = temp_end_research.CRA_name
    end_research.CRA_phoneNumber = temp_end_research.CRA_phoneNumber
    end_research.save()

    # 연구 바인더 위치 (사진)
    if request.FILES.getlist('image[]'):
        prev_images = End_UploadImage.objects.filter(is_deleted=False, end_research=end_research)
        prev_images.update(is_deleted=False)

    for _image in request.FILES.getlist('image[]'):
        new_image = End_UploadImage(imagename=_image.name, image=_image, end_research=end_research)
        new_image.save()

    # 장기 보관 문서 (파일)
    if request.FILES.getlist('end_research_archive[]'):
        prev_files = End_Research_Archive.objects.filter(is_deleted=False, end_research=end_research)
        prev_files.update(is_deleted=True)

    for _file in request.FILES.getlist('end_research_archive[]'):
        new_file = End_Research_Archive(filename=_file.name, file=_file, end_research=end_research)
        new_file.save()

    if len(request.FILES.getlist('end_research_archive[]')) > 0:
        binder_image = End_UploadImage.objects.filter(is_deleted=False, end_research=end_research)
        binder_image.update(is_deleted=True)

    return HttpResponseRedirect('/research/end_study/')

@login_required()
def delete_end_research(request, id):
    end_research = End_research.objects.get(pk=id)
    end_research.is_deleted = True
    end_research.save()
    return HttpResponseRedirect(f'/research/end_study/')

@login_required()
def search_end_study(request):
    if not request.GET or 'cancer' not in request.GET:
        all_end_study, query_dict = generate_end_study_search_query(request)

        return render(
            request, 'pages/research/end_study/end_study.html', {
                'field_choice': Research.create_field_value_and_text(),
                'editable': True,
                'query': query_dict,
                'search': all_end_study,
                'field_key_value': Research.create_field_value_and_text_dict(),
            }
        )


@login_required()
def download_statistics(request):
    statistics_download = DownloadLog(downloader=request.user, content=request.POST.get('content'))
    statistics_download.save()

    return HttpResponse()
